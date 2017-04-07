#!/usr/bin/env python
#-*- coding: utf-8 -*-
import json
from decimal import Decimal
from datetime import datetime

from django.core.exceptions import ObjectDoesNotExist
from django.shortcuts import render
from django.shortcuts import render_to_response
from django.template import RequestContext, TemplateDoesNotExist
from django.conf import settings
from django.core import serializers
from django.core.serializers.json import DjangoJSONEncoder
from django.db.models import Sum
from django.http import HttpResponse, HttpResponseRedirect, Http404
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.core.urlresolvers import reverse_lazy
from django.views.generic import ListView, TemplateView, View
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from openpyxl import load_workbook, Workbook, cell
from openpyxl.utils import coordinate_from_string, column_index_from_string, get_column_letter
from openpyxl.styles import Alignment, Style, Border, Side, colors, PatternFill, Color

from ..home.models import *
from ..rrhh.models import *
from ..ventas.models import Proyecto
from .forms import *
from .models import *
from ..tools.globalVariable import *

from ..tools.genkeys import *
from ..tools.uploadFiles import upload, deleteFile

class JSONResponseMixin(object):
    def render_to_json_response(self, context, **response_kwargs):
        return HttpResponse(
            self.convert_context_to_json(context),
            content_type='application/json',
            mimetype='application/json',
            **response_kwargs)

    def convert_context_to_json(self, context):
        return json.dumps(context, encoding='utf-8', cls=DjangoJSONEncoder)


'''
    Block for assistance
'''

class EmployeeHome(TemplateView):

    template_name = 'rrhh/index.html'

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        return render(request, self.template_name, kwargs)

class TypeEmployeeView(JSONResponseMixin, TemplateView):

    template_name = 'rrhh/typeemployee.html'

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        try:
            if request.is_ajax():
                try:
                    if 'ltypes' in request.GET:
                        kwargs['types'] = json.loads(
                            serializers.serialize(
                                'json',
                                TypesEmployee.objects.filter(flag=True).order_by('register')))
                        kwargs['status'] = True
                except ObjectDoesNotExist as oex:
                    kwargs['raise'] = str(oex)
                    kwargs['status'] = False
                return self.render_to_json_response(kwargs)
            return render(request, self.template_name, kwargs)
        except TemplateDoesNotExist as ext:
            raise Http404(ext)

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        if request.is_ajax():
            try:
                if 'new' in request.POST:
                    key = TypesEmployeeKeys()
                    TypesEmployee.objects.create(
                        types_id=key,
                        description=request.POST['desc'].upper(),
                        starthour=request.POST['starthour'],
                        outsaturday=request.POST['outsaturday'])
                    kwargs['status'] = True
                if 'modify' in request.POST:
                    obj = TypesEmployee.objects.get(types_id=request.POST['pk'])
                    obj.description = request.POST['desc'].upper()
                    obj.starthour = request.POST['starthour']
                    obj.outsaturday = request.POST['outsaturday']
                    obj.save()
                    kwargs['status'] = True
                if 'delete' in request.POST:
                    obj = TypesEmployee.objects.get(types_id=request.POST['pk'])
                    obj.delete()
                    kwargs['status'] = True
            except ObjectDoesNotExist as oex:
                kwargs['raise'] = str(oex)
                kwargs['status'] = False
            return self.render_to_json_response(kwargs)


# class for mantenice break employee
class EmployeeBreakView(JSONResponseMixin, TemplateView):
    template_name = 'rrhh/breakoemployee.html'

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        try:
            if request.is_ajax():
                try:
                    if 'getstatus' in request.GET:
                        kwargs['lstatus'] = json.loads(serializers.serialize(
                            'json',
                            EmployeeBreak.objects.filter(flag=True)))
                        kwargs['status'] = True
                except ObjectDoesNotExist as oex:
                    kwargs['raise'] = str(oex)
                    kwargs['status'] = False
                return self.render_to_json_response(kwargs)
            return render(request, self.template_name, kwargs)
        except TemplateDoesNotExist as extv:
            raise Http404(extv)

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        if request.is_ajax():
            try:
                if 'savestatus' in request.POST:
                    obj = None
                    if 'pk' in request.POST:
                        obj = EmployeeBreak.objects.get(
                            status_id=request.POST['pk'])
                    else:
                        obj = EmployeeBreak()
                        obj.status_id = StatusAssistanceId()
                    obj.description = str(request.POST['description']).upper()
                    obj.payment = request.POST['payment'] in (True, 'true', 'True')
                    obj.save()
                    kwargs['status'] = True
                if 'delete' in request.POST:
                    ebs = EmployeeBreak.objects.get(status_id=request.POST['pk'])
                    try:
                        ebs.delete()
                    except Exception as eex:
                        ebs.flag = False
                        ebs.save()
                        kwargs['msg'] = 'El estado no se ha eliminado %s' % str(eex)
                    kwargs['status'] = True
            except ObjectDoesNotExist as oex:
                kwargs['raise'] = str(oex)
                kwargs['status'] = False
            return self.render_to_json_response(kwargs)


# this class register assistenace of employee
class AssistanceEmployee(JSONResponseMixin, TemplateView):
    """this class implement asisstance employee"""
    template_name = 'rrhh/assistance.html'

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        try:
            if request.is_ajax():
                try:
                    if 'glist' in request.GET:
                        kwargs['employee'] = json.loads(
                            serializers.serialize(
                                'json',
                                Employee.objects.filter(flag=True),
                                relations=('charge')))
                        kwargs['status'] = True
                    if 'gproject' in request.GET:
                        kwargs['projects'] = json.loads(
                            serializers.serialize(
                                'json',
                                Proyecto.objects.filter(status='AC', flag=True)))
                        kwargs['status'] = True
                    if 'gettypes' in request.GET:
                        kwargs['types'] = json.loads(
                            serializers.serialize(
                                'json',
                                TypesEmployee.objects.filter(flag=True).order_by('description')))
                        kwargs['status'] = True
                except ObjectDoesNotExist as oex:
                    kwargs['raise'] = str(oex)
                    kwargs['status'] = False
                return self.render_to_json_response(kwargs)
            return  render(request, self.template_name, kwargs)
        except TemplateDoesNotExist as ext:
            raise Http404(ext)

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        if request.is_ajax():
            try:
                if 'saveAssistance' in request.POST:
                    # print request.POST
                    proj = request.POST['project'] if 'project' in request.POST else ''
                    proj = '' if proj == 'null' else proj
                    def registerassistance():
                        """this function register assistance"""
                        obj = Assistance()
                        obj.userregister_id = request.user.get_profile().empdni_id
                        obj.employee_id = request.POST['dni']
                        obj.project_id = proj if proj != '' else None
                        obj.types_id = request.POST['type']
                        obj.assistance = request.POST['date']
                        obj.hourin = request.POST['hin']
                        obj.hourout = request.POST['hout']
                        obj.hourinbreak = request.POST['hinb']
                        obj.houroutbreak = request.POST['houtb']
                        obj.viatical = request.POST['viatical']
                        obj.tag = True
                        obj.save()
                        return 1
                    exists = None
                    if proj != '' or proj is None:
                        exists = Assistance.objects.filter(
                            assistance=request.POST['date'],
                            employee_id=request.POST['dni'], project_id=proj)
                    else:
                        exists = Assistance.objects.filter(
                            assistance=request.POST['date'], employee_id=request.POST['dni'])
                    if len(exists):
                        obj = exists.latest('register')
                        hourin = request.POST['hin']
                        # print 'OBJECT ', type(obj.hourin), obj.hourin
                        hourin = str(hourin)[:5]
                        if type(hourin) is str:
                            inh = datetime.datetime.strptime(hourin, '%H:%M').time()
                        else:
                            inh = hourin
                        print type(inh), inh
                        print type(hourin), hourin
                        if inh <= obj.hourin:
                            kwargs['raise'] = 'No se puede por que la hora ingresada '\
                                'es menor a una ya ingresada para esta fecha'
                            kwargs['status'] = False
                        else:
                            registerassistance()
                    else:
                        registerassistance()
                    if 'status' not in kwargs:
                        kwargs['status'] = True
            except ObjectDoesNotExist as oex:
                kwargs['status'] = False
                kwargs['raise'] = str(oex)
            return self.render_to_json_response(kwargs)


# mantenice settings asisstance
class EmployeeAsisstanceView(JSONResponseMixin, TemplateView):

    template_name = 'rrhh/mantenice.html'

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        try:
            if request.is_ajax():
                try:
                    if 'gettypes' in request.GET:
                        kwargs['types'] = json.loads(
                            serializers.serialize(
                                'json',
                                TypesEmployee.objects.filter()
                            )
                        )
                        kwargs['status'] = True
                    if 'getsettings' in request.GET:
                        kwargs['settings'] = json.loads(
                            serializers.serialize(
                                'json',
                                EmployeeSettings.objects.filter(flag=True)
                            )
                        )
                        kwargs['status'] = True
                except ObjectDoesNotExist as exo:
                    kwargs['raise'] = str(exo)
                    kwargs['status'] = False
                return self.render_to_json_response(kwargs)
            return render(request, self.template_name, kwargs)
        except TemplateDoesNotExist as ext:
            raise Http404(ext)

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        # print request.is_ajax()
        if request.is_ajax():
            try:
                print request.POST
                if 'savesettings' in request.POST:
                    EmployeeSettings.objects.filter(flag=True).update(flag=False)
                    # print 'before declare'
                    emp = EmployeeSettings()
                    # print 'declare'
                    # print request.POST['hextperfirst'], type(request.POST['hextperfirst'])
                    emp.hextperfirst = Decimal(
                        request.POST['hextperfirst']).quantize(Decimal('0.1'))
                    emp.hextpersecond = Decimal(
                        request.POST['hextpersecond']).quantize(Decimal('0.1'))
                    emp.ngratification = request.POST['ngratification']
                    emp.ncts = request.POST['ncts']
                    emp.pergratification = Decimal(
                        request.POST['pergratification']).quantize(Decimal('0.1'))
                    emp.codeproject = request.POST['codeproject']
                    emp.starthourextra = request.POST['starthourextra']
                    emp.starthourextratwo = request.POST['starthourextratwo']
                    emp.shxsaturday = request.POST['shxsaturday']
                    emp.shxsaturdayt = request.POST['shxsaturdayt']
                    emp.totalhour = request.POST['totalhour']
                    emp.timeround = request.POST['timeround']
                    emp.save()
                    kwargs['status'] = True
            except ObjectDoesNotExist as oex:
                kwargs['raise'] = str(oex)
                kwargs['status'] = False
            return self.render_to_json_response(kwargs)


# load file for Assistance
class LoadAssistance(JSONResponseMixin, TemplateView):
    """ upload files for read and load data for the assistance """
    template_name = 'rrhh/loadfile.html'

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        try:
            if request.is_ajax():
                try:
                    pass
                except ObjectDoesNotExist as oex:
                    kwargs['raise'] = str(oex)
                    kwargs['status'] = False
                return self.render_to_json_response(kwargs)
            return render(request, self.template_name, kwargs)
        except TemplateDoesNotExist as ext:
            raise Http404(ext)

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        """ upload file """
        if request.is_ajax():
            try:
                if 'loadfiles' in request.POST:
                    # print 'Files ', request.FILES
                    valid = True
                    counter = 0
                    # load and validate format file
                    try:
                        # upload file in the path
                        path = '\\storage\\assistance\\'
                        filename = upload(path, request.FILES['files'])
                        raises = list()
                        print filename
                        # validate file is meets with format
                        wbook = load_workbook(filename=filename, data_only=True)
                        if 'Asistencia' not in wbook.get_sheet_names():
                            kwargs['raise'] = 'La hoja no existe en el archivo'
                            valid = False
                        else:
                            wsheet = wbook['Asistencia']
                            nrow = wsheet.max_row
                            ncol = wsheet.max_column
                            print 'ROW ', nrow
                            print 'COLUMN ', ncol
                            if nrow < 9:
                                valid = False
                                kwargs['raise'] = 'El archivo no cumple con el '\
                                    'numero minimo filas para registrar'
                            if ncol < 44:
                                kwargs['raise'] = 'El archivo no cumple con el numero'\
                                    ' minimo de columnas para registrar'
                                valid = False
                            if valid:
                                param = {}
                                # get code project  for employee settings
                                sett = EmployeeSettings.objects.filter(flag=True)[0]
                                def registerassistance(parameter):
                                    """this function register assistance"""
                                    try:
                                        obj = Assistance()
                                        obj.userregister_id = request.user.get_profile().empdni_id
                                        obj.employee_id = parameter['employee']
                                        obj.project_id = parameter['project']
                                        obj.types_id = parameter['types']
                                        obj.assistance = parameter['assistance']
                                        obj.hourin = parameter['hourin']
                                        obj.hourout = parameter['hourout']
                                        obj.hourinbreak = parameter['hourinbreak']
                                        obj.houroutbreak = parameter['houroutbreak']
                                        obj.viatical = parameter['viatical']
                                        obj.tag = True
                                        obj.save()
                                        print 'OK'
                                        return 1
                                    except Exception as oex:
                                        raises.append({
                                            'employee': parameter['employee'],
                                            'assistance': parameter['assistance'],
                                            'types': parameter['types'],
                                            'raise': str(oex)})
                                        return 0
                                keys = {
                                    1: 'hourin',
                                    2: 'hourinbreak',
                                    3: 'houroutbreak',
                                    4: 'hourout',
                                    5: 'viatical',
                                    6: 'project'}
                                def validformat(oparam):
                                    """ Valid format object in and fields null """
                                    print 'inside valid format '
                                    if oparam['hourinbreak'] is None:
                                        oparam['hourinbreak'] = '00:00'
                                    else:
                                        # print type(oparam['hourinbreak'])
                                        if type(oparam['hourinbreak']) is datetime:
                                            oparam['hourinbreak'] = oparam['hourinbreak'].time()
                                    if oparam['houroutbreak'] is None:
                                        oparam['houroutbreak'] = '00:00'
                                    else:
                                        if type(oparam['houroutbreak']) is datetime:
                                            oparam['houroutbreak'] = oparam['houroutbreak'].time()
                                    if oparam['hourout'] is not None:
                                        if type(oparam['hourout']) is datetime:
                                            oparam['hourout'] = oparam['hourout'].time()
                                    if oparam['hourin'] is not None:
                                        if type(oparam['hourin']) is datetime:
                                            oparam['hourin'] = oparam['hourin'].time()
                                    if oparam['viatical'] is None:
                                        oparam['viatical'] = 0
                                    print 'end format'
                                    return oparam
                                for xraw in xrange(9, nrow+1):
                                    param = {}
                                    # ndta = 3
                                    dni = wsheet.cell(row=xraw, column=2).internal_value
                                    if dni is None:
                                        continue
                                    dni = str(dni)
                                    print dni
                                    if len(dni) != 8:
                                        continue
                                    nreg = 1
                                    xcol = 2
                                    while xcol <= 44:
                                        xcol += 1
                                        if xcol > 44:
                                            break
                                        param['employee'] = dni
                                        try:
                                            if xcol % 3 == 0:
                                                assistance = wsheet.cell(row=7, column=xcol).value
                                                if assistance is not None:
                                                    print assistance
                                                    param['assistance'] = assistance
                                                else:
                                                    if nreg == 1:
                                                        xcol += 5
                                                        param = {}
                                                        continue
                                            cval = wsheet.cell(row=xraw, column=xcol).value
                                            # print cval
                                            param[keys[nreg]] = cval
                                            if nreg == 1:
                                                if param['hourin'] is None:
                                                    xcol += 5
                                                    param = {}
                                                    continue
                                            if nreg == 6:
                                                npro = ('project' in param
                                                        and param['project'] is not None)
                                                if npro:
                                                    if len(param['project']) == 7:
                                                        param['types'] = sett.codeproject
                                                    else:
                                                        if len(param['project']) == 4:
                                                            param['types'] = param['project']
                                                            param['project'] = None
                                                        else:
                                                            param['project'] = None
                                                            param['types'] = 'TY04'
                                                else:
                                                    if param['project'] is None:
                                                        param['types'] = 'TY04'
                                                    else:
                                                        if len(param['project']) == 4:
                                                            param['types'] = param['project']
                                                            param['project'] = None
                                                        else:
                                                            param['types'] = 'TY04'
                                                            param['project'] = None
                                                if param['hourin'] is not None:
                                                    nreg = 1
                                                    param = validformat(param)
                                                    print param
                                                    counter += registerassistance(param)
                                                    param = {}
                                            else:
                                                nreg += 1
                                        except Exception as exp:
                                            print exp
                                    print '============================================='
                        kwargs['status'] = valid
                    except Exception as exio:
                        kwargs['raise'] = str(exio)
                        kwargs['status'] = False
                    wbook.save(filename)
                    # time.sleep(30)
                    deleteFile(filename)
                    kwargs['raises'] = raises
                    print counter
            except ObjectDoesNotExist as oex:
                kwargs['status'] = False
                kwargs['raise'] = str(oex)
        return self.render_to_json_response(kwargs)


# Class for view assistance loaded
class ViewAssistance(JSONResponseMixin, TemplateView):

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        try:
            if request.is_ajax():
                try:
                    if 'projects' in request.GET:
                        kwargs['projects'] = json.loads(serializers.serialize(
                            'json',
                            Proyecto.objects.filter(
                                flag=True, status='AC').order_by('-registrado')))
                        kwargs['status'] = True
                    if 'gtypes' in request.GET:
                        kwargs['types'] = json.loads(serializers.serialize(
                            'json',
                            TipoEmpleado.objects.filter(flag=True)))
                        kwargs['status'] = True
                    if 'lrtypes' in request.GET:
                        kwargs['ltypes'] = json.loads(serializers.serialize(
                            'json',
                            TypesEmployee.objects.filter(flag=True)))
                        kwargs['status'] = True
                    if 'getday' in request.GET:
                        kwargs['hours'] = json.loads(serializers.serialize(
                            'json',
                            Assistance.objects.filter(
                                employee_id=request.GET['dni'],
                                assistance=request.GET['day']).order_by('hourin'),
                            relations=('project', 'types', 'status')))
                        kwargs['status'] = True
                    if 'filterAssistance' in request.GET:
                        if 'type' in request.GET:
                            ass = BalanceAssistance.objects.filter(
                                assistance__range=(
                                    request.GET['weekstart'],
                                    request.GET['weekend']),
                                employee__tipoemple_id=request.GET['type']).order_by('assistance')
                        else:
                            ass = BalanceAssistance.objects.filter(
                                assistance__range=(
                                    request.GET['weekstart'],
                                    request.GET['weekend'])).order_by('assistance')
                        week = []
                        for x in ass:
                            count = 0
                            index = 0
                            for sw in xrange(0, len(week)):
                                if week[sw]['dni'] != x.employee_id:
                                    count += 1
                                    continue
                                else:
                                    index = sw
                                    break
                            if count == len(week):
                                week.append({
                                    'dni': x.employee_id,
                                    'name': '%s %s' % (x.employee.lastname, x.employee.firstname),
                                    'days': {x.assistance.strftime('%A'): {
                                        'day': x.assistance,
                                        'hour': x.hwork,
                                        'extra': (x.hextfirst + x.hextsecond),
                                        'delay': x.hdelay,
                                        'lack': x.hlack}},
                                    'count': 1,
                                    'twork': x.hwork,
                                    'textra': (x.hextfirst + x.hextsecond),
                                    'discount': x.discount})
                            else:
                                week[index]['days'][x.assistance.strftime('%A')] = {
                                    'day': x.assistance,
                                    'hour': x.hwork,
                                    'extra': (x.hextfirst + x.hextsecond),
                                    'delay': x.hdelay,
                                    'lack': x.hlack}
                                week[index]['count'] += 1
                                week[index]['twork'] += x.hwork
                                week[index]['textra'] += (x.hextfirst + x.hextsecond)
                                week[index]['discount'] += x.discount
                        nweek = [
                            'Monday',
                            'Tuesday',
                            'Wednesday',
                            'Thursday',
                            'Friday',
                            'Saturday',
                            'Sunday']
                        for x in week:
                            for w in nweek:
                                if w not in x['days']:
                                    x['days'][w] = {'day': None}
                                else:
                                    x['viatical'] = Assistance.objects.filter(
                                        assistance=x['days'][w]['day'],
                                        employee_id=x['dni']).aggregate(
                                            Sum('viatical'))['viatical__sum']
                        kwargs['week'] = week
                        name = {}
                        day = datetime.datetime.strptime(request.GET['weekstart'], '%Y-%m-%d')
                        traslate = {
                            'Monday': 'Lunes',
                            'Tuesday': 'Martes',
                            'Wednesday': 'Miercoles',
                            'Thursday': 'Jueves',
                            'Friday': 'Viernes',
                            'Saturday': 'Sabado',
                            'Sunday': 'Domingo'}
                        for x in xrange(0, 7):
                            name[x] = {
                                'nmo': day.strftime('%A'),
                                'nmt': traslate[day.strftime('%A')],
                                'nm': day.strftime('%m/%d'),
                                'date': day.strftime('%Y-%m-%d')}
                            day = day + datetime.timedelta(days=1)
                        conf = EmployeeSettings.objects.get(flag=True)
                        kwargs['thour'] = (conf.totalhour.hour) + ((conf.totalhour.minute)/60.0)
                        kwargs['names'] = name
                        kwargs['status'] = True
                except (ObjectDoesNotExist or Exception) as oex:
                    kwargs['raise'] = str(oex)
                    kwargs['status'] = False
                return self.render_to_json_response(kwargs)
            return render(request, 'rrhh/viewassistance.html', kwargs)
        except TemplateDoesNotExist as ext:
            return Http404(ext)

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        if request.is_ajax():
            try:
                if 'saveAssistance' in request.POST:
                    if 'modifyassistance' in request.POST:
                        asl = Assistance.objects.filter(
                            employee_id=request.POST['dni'], assistance=request.POST['date'])
                        if asl:
                            try:
                                print "ready for delete assistance"
                                asl.get(pk=request.POST['pkassistance']).delete()
                                if len(asl) == 1:
                                    BalanceAssistance.objects.get(
                                        employee_id=request.POST['dni'],
                                        assistance=request.POST['date']).delete()
                            except Exception as bex:
                                pass
                    proj = request.POST['project'] if 'project' in request.POST else ''
                    proj = '' if proj == 'null' else proj
                    def registerassistance():
                        """this function register assistance"""
                        obj = Assistance()
                        obj.userregister_id = request.user.get_profile().empdni_id
                        obj.employee_id = request.POST['dni']
                        obj.project_id = proj if proj != '' else None
                        obj.types_id = request.POST['type']
                        obj.assistance = request.POST['date']
                        obj.hourin = request.POST['hin']
                        obj.hourout = request.POST['hout']
                        obj.hourinbreak = request.POST['hinb']
                        obj.houroutbreak = request.POST['houtb']
                        obj.viatical = request.POST['viatical']
                        obj.tag = True
                        obj.save()
                        return 1
                    exists = None
                    if proj != '' or proj is None:
                        exists = Assistance.objects.filter(
                            assistance=request.POST['date'],
                            employee_id=request.POST['dni'], project_id=proj)
                    else:
                        exists = Assistance.objects.filter(
                            assistance=request.POST['date'], employee_id=request.POST['dni'])
                    if len(exists):
                        obj = exists.latest('register')
                        hourin = request.POST['hin']
                        # print 'OBJECT ', type(obj.hourin), obj.hourin
                        print type(hourin), hourin
                        hourin = str(hourin)[:5]
                        if type(hourin) is str:
                            inh = datetime.datetime.strptime(hourin, '%H:%M').time()
                            # inh = datetime.strptime(hourin, '%H:%M').time()
                        else:
                            inh = hourin
                        print type(inh), inh
                        print type(hourin), hourin
                        if inh <= obj.hourin:
                            kwargs['raise'] = 'No se puede por que la hora ingresada '\
                                'es menor a una ya ingresada para esta fecha'
                            kwargs['status'] = False
                        else:
                            registerassistance()
                    else:
                        registerassistance()
                    if 'status' not in request.POST:
                        try:
                            rb = BalanceAssistance.objects.get(
                                employee_id=request.POST['dni'],
                                assistance=request.POST['date'])
                            rb.discount = request.POST['discount']
                            rb.save()
                        except Exception as rdex:
                            pass
                    kwargs['status'] = True
                if 'deleteAssistance' in request.POST:
                    asl = Assistance.objects.filter(
                        employee_id=request.POST['dni'], assistance=request.POST['date'])
                    if asl:
                        try:
                            asl.get(pk=request.POST['pk']).delete()
                            if len(asl) == 1:
                                BalanceAssistance.objects.get(
                                    employee_id=request.POST['dni'],
                                    assistance=request.POST['date']).delete()
                            else:
                                asl = Assistance.objects.filter(
                                    employee_id=request.POST['dni'],
                                    assistance=request.POST['date'])
                                asl[0].viatical = asl[0].viatical
                                asl[0].save()
                        except Exception as bex:
                            pass
                        kwargs['status'] = True
            except ObjectDoesNotExist as oex:
                kwargs['status'] = False
                kwargs['raise'] = str(oex)
            return self.render_to_json_response(kwargs)


class ExportarAssistance(JSONResponseMixin, TemplateView):

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        try:
            if request.is_ajax():
                if 'gettypes' in request.GET:
                    kwargs['types'] = json.loads(serializers.serialize(
                        'json',
                        TipoEmpleado.objects.filter(flag=True)))
                    kwargs['status'] = True
                if 'getpayment' in request.GET:
                    kwargs['payment'] = json.loads(serializers.serialize(
                        'json',
                        TipoPago.objects.filter(flag=True)))
                    kwargs['status'] = True
                return self.render_to_json_response(kwargs)
            print request.GET
            if 'exportdata' in request.GET:
                kwargs['status'] = True
                if 'type' in request.GET:
                    ass = BalanceAssistance.objects.filter(
                        assistance__range=(
                            request.GET['din'],
                            request.GET['dout']),
                        employee__tipoemple_id=request.GET['type']).order_by('employee__lastname')
                else:
                    ass = BalanceAssistance.objects.filter(
                        assistance__range=(
                            request.GET['din'],
                            request.GET['dout'])).order_by('employee__lastname')
                week = []
                for x in ass:
                    count = 0
                    index = 0
                    for sw in xrange(0, len(week)):
                        if week[sw]['dni'] != x.employee_id:
                            count += 1
                            continue
                        else:
                            index = sw
                            break
                    if count == len(week):
                        print x.employee_id
                        repen = detPensEmple.objects.get(empdni_id=x.employee_id)
                        cuenta = CuentaEmple.objects.filter(
                            empdni_id=x.employee_id).order_by('-registro')[0]
                        week.append({
                            'dni': x.employee_id,
                            'name': '%s %s' % (x.employee.lastname, x.employee.firstname),
                            'pensionario': repen.regimenpens.regimen,
                            'percent': repen.regimenpens.percent,
                            'cuenta': cuenta.cuenta,
                            'remuneracion': cuenta.remuneracion,
                            'setfamily': cuenta.setfamily,
                            # 'days': {x.assistance.strftime('%A'): {
                            #     'day': x.assistance,
                            #     'hour': x.hwork,
                            #     'extra': (x.hextfirst + x.hextsecond),
                            #     'delay': x.hdelay,
                            #     'lack': x.hlack}},
                            'count': 1,
                            'twork': x.hwork,
                            'textra': (x.hextfirst + x.hextsecond),
                            'discount': x.discount,
                            'delay': x.hdelay,
                            'viatical': Assistance.objects.filter(
                                assistance__range=(request.GET['din'], request.GET['dout']),
                                employee_id=x.employee_id).aggregate(
                                    Sum('viatical'))['viatical__sum']})
                    else:
                        # week[index]['days'][x.assistance.strftime('%A')] = {
                        #     'day': x.assistance,
                        #     'hour': x.hwork,
                        #     'extra': (x.hextfirst + x.hextsecond),
                        #     'delay': x.hdelay,
                        #     'lack': x.hlack}
                        week[index]['count'] += 1
                        week[index]['twork'] += x.hwork
                        week[index]['textra'] += (x.hextfirst + x.hextsecond)
                        week[index]['discount'] += x.discount
                        week[index]['delay'] += x.hdelay
                # Create File ExportarAssistance
                # block start style
                def style_range(ws, cell_range, options={}):
                    """
                    :param ws:  Excel worksheet instance
                    :param range: An excel range to border (e.g. A1:F20)
                    :param border: An openpyxl border object
                    """

                    start_cell, end_cell = cell_range.split(':')
                    start_coord = coordinate_from_string(start_cell)
                    start_row = start_coord[1]
                    start_col = column_index_from_string(start_coord[0])
                    end_coord = coordinate_from_string(end_cell)
                    end_row = end_coord[1]
                    end_col = column_index_from_string(end_coord[0])
                    for row in range(start_row, end_row + 1):
                        for col_idx in range(start_col, end_col + 1):
                            col = get_column_letter(col_idx)
                            if 'fnumber' in options:
                                ws.cell('%s%s' % (col, row)).number_format = options['fnumber']
                            if 'border' in options:
                                ws.cell('%s%s' % (col, row)).border = options['border']
                            if 'bg' in options:
                                ws.cell('%s%s' % (col, row)).fill = options['bg']
                            if 'alignment' in options:
                                ws.cell('%s%s' % (col, row)).alignment = options['alignment']
                # end block
                tf = {'hour': 'HORA', 'days': 'DIAS'}
                response = HttpResponse(
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment;filename=PLANILLA-%s-%s X %s.xlsx' % (
                    request.GET['din'], request.GET['dout'], tf[request.GET['tfile']])
                wb = Workbook()
                ws = wb.active
                ws.title = 'PANILLA %s AL %s' % (request.GET['din'][5:], request.GET['dout'][5:])
                ws.sheet_properties.tabColor = '1072BA'
                # set title
                ws.cell(column=1, row=3).value = 'PLANILLA DE %s AL %s' % (
                    request.GET['din'], request.GET['dout'])
                #set Header
                ws.cell(column=2, row=4).value = 'Apellidos y Nombres'
                ws.cell(column=3, row=4).value = 'Sistema de Pensiones'
                ws.cell(column=4, row=4).value = '%'
                ws.cell(column=5, row=4).value = 'Jornal'
                ws.cell(column=6, row=4).value = 'Dominical'
                ws.cell(column=7, row=4).value = 'Asig. Familiar'
                ws.cell(column=8, row=4).value = 'Total Rem.'
                ws.cell(column=9, row=4).value = '%s Laboradas' % (tf[request.GET['tfile']])
                ws.cell(column=10, row=4).value = 'Horas Extra'
                ws.cell(column=11, row=4).value = 'Rem. Total'
                ws.cell(column=12, row=4).value = 'AFP/ONP'
                ws.cell(column=13, row=4).value = 'Otros Dsct'
                ws.cell(column=14, row=4).value = 'Total Dsct'
                ws.cell(column=15, row=4).value = 'Total a Pagar'
                # set data body
                count = 5
                total = 0
                reminv = 85
                pxhour = 8
                print week
                print len(week)
                if request.GET['tfile'] == 'hour':
                    for x in week:
                        dom = 0
                        sfamily = 0
                        dom = Decimal(x['remuneracion'] / float(
                            request.GET['payment'])).quantize(Decimal('0.01'))
                        if x['setfamily'] == True:
                            sfamily = Decimal((float(reminv)/30) * float(
                                request.GET['payment'])).quantize(Decimal('0.01'))
                        ws.cell(column=1, row=count).value = (count - 4)
                        ws.cell(column=2, row=count).value = x['name']
                        ws.cell(column=3, row=count).value = x['pensionario']
                        ws.cell(column=4, row=count).value = Decimal(
                            x['percent']).quantize(Decimal('0.01'))
                        ws.cell(column=5, row=count).value = (
                            (Decimal(x['remuneracion']) - dom) - sfamily) if x['setfamily'] else (
                                Decimal(x['remuneracion']) - dom)
                        ws.cell(column=6, row=count).value = dom
                        ws.cell(column=7, row=count).value = sfamily
                        ws.cell(column=8, row=count).value = x['remuneracion']
                        ## to here performed calc hours
                        calchour = Decimal((
                            x['remuneracion'] / float(request.GET['payment'])) / pxhour).quantize(
                                Decimal('0.0001'))
                        ws.cell(column=9, row=count).value = x['twork']
                        ws.cell(column=10, row=count).value = x['textra']
                        tworked = Decimal(calchour * x['twork']).quantize(Decimal('0.01'))
                        ws.cell(column=11, row=count).value = tworked + dom
                        safp = Decimal(
                            float(tworked + dom) * (float(x['percent']) / 100)).quantize(Decimal('0.01'))
                        ws.cell(column=12, row=count).value = safp
                        ws.cell(column=13, row=count).value = x['discount']
                        ws.cell(column=14, row=count).value = (x['discount'] + safp)
                        tot = ((tworked + dom) - (safp))
                        ws.cell(column=15, row=count).value = tot
                        total += tot
                        count += 1
                    ws.cell(column=15, row=count).value = total
                if request.GET['tfile'] == 'days':
                    for x in week:
                        dom = 0
                        sfamily = 0
                        dom = Decimal(x['remuneracion'] / float(
                            request.GET['payment'])).quantize(Decimal('0.01'))
                        if x['setfamily'] == True:
                            sfamily = Decimal((float(reminv)/30) * float(request.GET['payment'])).quantize(Decimal('0.01'))
                        ws.cell(column=1, row=count).value = (count - 4)
                        ws.cell(column=2, row=count).value = x['name']
                        ws.cell(column=3, row=count).value = x['pensionario']
                        ws.cell(column=4, row=count).value = Decimal(
                            x['percent']).quantize(Decimal('0.01'))
                        ws.cell(column=5, row=count).value = (
                            (Decimal(x['remuneracion']) - dom) - sfamily) if x['setfamily'] else (
                                Decimal(x['remuneracion']) - dom)
                        ws.cell(column=6, row=count).value = dom
                        ws.cell(column=7, row=count).value = sfamily
                        ws.cell(column=8, row=count).value = x['remuneracion']
                        ## to here performed calc hours
                        calchour = Decimal((
                            x['remuneracion'] / float(request.GET['payment'])) / pxhour).quantize(
                                Decimal('0.0001'))
                        calcday = Decimal((
                            float(x['remuneracion']) / float(request.GET['payment']))).quantize(
                                Decimal('0.0001'))
                        ws.cell(column=9, row=count).value = x['count']
                        ws.cell(column=10, row=count).value = x['textra']
                        tworked = Decimal((x['count'] * calcday)).quantize(Decimal('0.01'))
                        ws.cell(column=11, row=count).value = tworked + dom
                        safp = Decimal(
                            float(tworked + dom) * (float(x['percent']) / 100)).quantize(Decimal('0.01'))
                        ws.cell(column=12, row=count).value = safp
                        ws.cell(column=13, row=count).value = x['discount']
                        ws.cell(column=14, row=count).value = (x['discount'] + safp)
                        tot = ((tworked + dom) - (safp))
                        ws.cell(column=15, row=count).value = tot
                        total += tot
                        count += 1
                    ws.cell(column=15, row=count).value = total
                style_range(ws, ('A3:O%d' % count), {
                    'border': Border(top=Side(border_style='thin', color=colors.BLACK),
                                    left=Side(border_style='thin', color=colors.BLACK),
                                    bottom=Side(border_style='thin', color=colors.BLACK),
                                    right=Side(border_style='thin', color=colors.BLACK), )},)
                style_range(ws, 'A3:O4', {'bg': PatternFill(
                    start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')})
                style_range(ws, ('O5:O%d' % count), {'bg': PatternFill(
                    start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')})
                ws['A3'].alignment = Alignment(horizontal='center')
                ws.merge_cells('A3:O3')
                ws.merge_cells('A%(n)d:N%(n)d' % {'n':count})
                ws['A%d'%count].value = 'TOTAL'
                ws['A%d'%count].alignment = Alignment(horizontal='right')
                ws.column_dimensions['A'].width = 4
                ws.column_dimensions['B'].width = 35
                ws.column_dimensions['C'].width = 12
                ws.column_dimensions['D'].width = 7
                style_range(ws, ('D5:O%d' % count), {'fnumber': '#,##0.00'})
                # now include employee with payment extra
                count += 3
                second = count
                ws.cell(column=1, row=count).value = 'EXTRAS'
                ws.merge_cells('A%(n)d:G%(n)d'%{'n': count})
                count += 1
                ws.cell(column=1, row=count).value = '#'
                ws.cell(column=2, row=count).value = 'APELLIDOS Y NOMBRES'
                ws.cell(column=3, row=count).value = 'HORAS EXTRAS'
                ws.cell(column=4, row=count).value = 'IMPORTE'
                ws.cell(column=5, row=count).value = 'VIATICOS'
                ws.cell(column=6, row=count).value = 'DSCT'
                ws.cell(column=7, row=count).value = 'TOTAL'
                count += 1
                nb = 1
                totalextra = 0
                tviatical = 0
                thours = 0
                tdiscount = 0
                style_range(ws, ('A%d:G%d' % (count - 2, count - 1)), {
                    'bg': PatternFill(
                        fill_type='solid', start_color='FFFFFF00', end_color='FFFFFF00')})
                for x in week:
                    ws.cell(column=1, row=count).value = (nb)
                    ws.cell(column=2, row=count).value = x['name']
                    ws.cell(column=3, row=count).value = x['textra']
                    calchour = Decimal((
                        x['remuneracion'] / float(request.GET['payment'])) / pxhour).quantize(
                            Decimal('0.0001'))
                    tsc = Decimal(calchour * x['textra']).quantize(Decimal('0.01'))
                    ws.cell(column=4, row=count).value = tsc
                    ws.cell(column=5, row=count).value = x['viatical']
                    ws.cell(column=6, row=count).value = x['discount']
                    ws.cell(column=7, row=count).value = Decimal(
                        (tsc + x['viatical']) - x['discount']).quantize(Decimal('0.01'))
                    totalextra += ws.cell(column=7, row=count).value
                    tviatical += x['viatical']
                    tdiscount += x['discount']
                    thours += x['textra']
                    nb += 1
                    count += 1
                ws.cell(column=3, row=count).value = thours
                ws.cell(column=5, row=count).value = tviatical
                ws.cell(column=6, row=count).value = tdiscount
                ws.cell(column=7, row=count).value = totalextra
                style_range(ws, ('A%d:G%d' % (second, count)), {
                    'border': Border(
                        top=Side(border_style='thin', color=colors.BLACK),
                        left=Side(border_style='thin', color=colors.BLACK),
                        bottom=Side(border_style='thin', color=colors.BLACK),
                        right=Side(border_style='thin', color=colors.BLACK), )})
                wb.save(response)
                return response
            else:
                return render(request, 'rrhh/exportar.html', kwargs)
        except TemplateDoesNotExist as ext:
            raise Http404(ext)

# class delete assistance
class DeleteAssistance(JSONResponseMixin, TemplateView):

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        try:
            kwargs['types'] = TipoEmpleado.objects.filter(flag=True)
            return render(request, 'rrhh/delassistance.html', kwargs)
        except TemplateDoesNotExist as ext:
            raise Http404(ext)

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        if  request.is_ajax():
            try:
                if 'deleteAssistance' in request.POST:
                    sass = Assistance.objects.filter(
                        assistance__range=(request.POST['start'], request.POST['end']),
                        employee__tipoemple_id=request.POST['type'])
                    for x in sass:
                        x.delete()
                    bass = BalanceAssistance.objects.filter(
                        assistance__range=(request.POST['start'], request.POST['end']),
                        employee__tipoemple_id=request.POST['type'])
                    for x in bass:
                        x.delete()
                    kwargs['status'] = True
            except ObjectDoesNotExist as oex:
                kwargs['status'] = False
                kwargs['raise'] = str(oex)
        return self.render_to_json_response(kwargs)

# #class load data test
class LoadRe(TemplateView):

    def get(self, request, *args, **kwargs):
        import  random
        # import date
        dni = ['10278772','09434564','43504371','47403858','70015965','09279015','19242694','71744452','43572763','44684519','45576052','09164147','43179067','40472915','17589085','41452883','23392439','47720073','10659544','42492201','40894066','43402155','45505146','42661096','00000000','46377134','77498386','44488336','44263241','46296730','42577451','43402156','46279300','47431405','44463959','42333571','72604244','70921414','77070893','70492850','46824555','43414194','42825437','41822143','44322016','40080344','05406738','47542298','43378121','70266206','43316178','44969405','42310463','43984365','40990712','40143637','43598455','40043764','45768794','47999575','71251518','43869036','46879331','48022378','73585896','25798318','44363536']
        count = 0
        if 'pensionary' in request.GET:
            afp = ['RP001', 'RP002', 'RP003', 'RP004', 'RP005']
            name = {'RP001':'HABITAT', 'RP002':'ONP', 'RP003':'PROFUTURO', 'RP004':'PRIMA', 'RP005':'INTEGRA'}
            percent = {'RP001':12.58, 'RP002':13.00, 'RP003': 12.53, 'RP004': 12.58, 'RP005': 12.53}
            count = 0
            for x in dni:
                try:
                    fc = datetime.datetime.today()
                    ds = detPensEmple()
                    ds.empdni_id = x
                    ds.regimenpens_id = afp[random.randrange(5)]
                    ds.fechinicio = fc
                    ds.fechfin = fc.replace(fc.year + 1)
                    ds.cuspp = '%s' % '{:0>20d}'.format(random.randint(10000, 100000))
                    ds.sctr = random.randrange(2)
                    ds.save()
                    count += 1
                except ObjectDoesNotExist as oex:
                    print oex
            kwargs['affected'] = 'LOAD REGIMENS'
        if 'accounts' in request.GET:
            account = None
            count = 0
            for x in dni:
                try:
                    CuentaEmple.objects.get(empdni_id=x)
                except ObjectDoesNotExist as oex:
                    account = '%i-%s-%i' % (random.randint(100, 999), '{:0>8d}'.format(random.randint(100, 9999)), random.randint(100, 999))
                    amount = random.randint(150, 500)
                    ce = CuentaEmple()
                    ce.empdni_id = x
                    ce.cuenta = account
                    ce.tipodepago_id = 'PA002'
                    ce.estado = 'ACTIVO'
                    ce.remuneracion = amount
                    ce.tipocontrato_id = 'CT001'
                    ce.cts = amount
                    ce.gratificacion = amount * 2
                    ce.costxhora = (amount / 7) / 8
                    ce.flag = True
                    ce.save()
                    count+=1
            kwargs['affected'] = 'LOAD ACCOUNTS'
        kwargs['transact'] = count
        return render(request, 'rrhh/test.html', kwargs)
