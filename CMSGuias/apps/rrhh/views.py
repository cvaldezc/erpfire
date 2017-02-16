#!/usr/bin/env python
#-*- coding: utf-8 -*-
import json
from decimal import Decimal
from datetime import datetime

from django.core.exceptions import ObjectDoesNotExist
from django.shortcuts import render
from django.shortcuts import render_to_response
from django.template import RequestContext, TemplateDoesNotExist
from django.conf import settings
from django.core import serializers
from django.core.serializers.json import DjangoJSONEncoder
from django.db.models import Sum
from django.http import HttpResponse, HttpResponseRedirect, Http404
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.core.urlresolvers import reverse_lazy
from django.views.generic import ListView, TemplateView, View
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from openpyxl import load_workbook, Workbook, cell
from openpyxl.utils import coordinate_from_string, column_index_from_string, get_column_letter
from openpyxl.styles import Alignment, Style, Border, Side, colors, PatternFill, Color

from ..home.models import *
from ..rrhh.models import *
from ..ventas.models import Proyecto
from .forms import *
from .models import *
from ..tools.globalVariable import *

from ..tools.genkeys import *
from ..tools.uploadFiles import upload, deleteFile


class JSONResponseMixin(object):
    def render_to_json_response(self, context, **response_kwargs):
        return HttpResponse(
            self.convert_context_to_json(context),
            content_type='application/json',
            mimetype='application/json',
            **response_kwargs)

    def convert_context_to_json(self, context):
        return json.dumps(context, encoding='utf-8', cls=DjangoJSONEncoder)


class EmployeeHome(TemplateView):

    template_name = 'rrhh/index.html'

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        return render(request, self.template_name, kwargs)


class EmpleDetails(ListView):

    template_name = 'rrhh/formviewdetempl.html'

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        context['listemple'] = Employee.objects.filter(empdni_id=kwargs['empdni_id'])
        context['listcuenta'] = CuentaEmple.objects.filter(empdni_id=kwargs['empdni_id']).order_by('estado')
        context['listphone'] = PhoneEmple.objects.filter(empdni_id=kwargs['empdni_id'])
        context['listsusp'] = detSuspension.objects.filter(empdni_id=kwargs['empdni_id']).order_by('-fechfin')
        context['listestudio'] = detEstudioEmple.objects.filter(empdni_id=kwargs['empdni_id']).order_by('-anoEgreso')
        context['listexamenes'] = ExamenEmple.objects.filter(empdni_id=kwargs['empdni_id']).order_by('-fechcaduca')
        context['listdocumentos'] = DocumentoEmple.objects.filter(empdni_id=kwargs['empdni_id']).order_by('-fechcaduca')
        context['listepps'] = Epps.objects.filter(empdni_id=kwargs['empdni_id']).order_by('item')
        context['listfam'] = FamiliaEmple.objects.filter(empdni_id=kwargs['empdni_id']).order_by('nombres')
        context['listexplab'] = ExperienciaLab.objects.filter(empdni_id=kwargs['empdni_id'])
        context['listemerg'] = CasoEmergencia.objects.filter(empdni_id=kwargs['empdni_id']).order_by('nombres')
        context['listfamicr'] = FamiliaIcr.objects.filter(empdni_id=kwargs['empdni_id']).order_by('nombres')
        context['listmedic'] = DatoMedico.objects.filter(empdni_id=kwargs['empdni_id'])
        return render_to_response(
            self.template_name,
            context,
            context_instance=RequestContext(request))

class EmpleNoActive(JSONResponseMixin, TemplateView):

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        emple = RenunciaEmple.objects.filter(empdni__flag=False).order_by('-ultimodiatrab')
        return render(request, 'rrhh/noactive.html', {'lnoactive': emple})


class Menu(JSONResponseMixin, TemplateView):
    """docstring for Menu"""

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        if request.is_ajax():
            try:
                if 'showeditemp' in request.GET:
                    emp = Employee.objects.values(
                        'empdni_id',
                        'firstname',
                        'lastname',
                        'birth',
                        'address',
                        'charge_id',
                        'charge_id__cargos',
                        'tipoemple_id',
                        'tipoemple_id__descripcion',
                        'sexo',
                        'estadocivil',
                        'address2',
                        'nacionalidad',
                        'discapacidad',
                        'feching',
                        'archivo',
                        'foto',
                        'estadoplanilla',
                        'nacdpt',
                        'nacprov',
                        'nacdist',
                        'tallazap',
                        'tallapolo',
                        'cargopostula',
                        'distrito',
                        'email').get(empdni_id=request.GET.get('employeeid'))
                    if len(emp) > 0:
                        emp['status'] = True

                dicEmple = {}
                for x, num in emp.items():
                    dicEmple[x] = num

                context['status'] = True
            except ObjectDoesNotExist, e:
                context['raise'] = str(e)
                context['status'] = False
            return self.render_to_json_response(dicEmple)

        emple = Employee.objects.filter(flag=True)
        tipemp = TipoEmpleado.objects.all()
        cargos = Cargo.objects.all()
        tipoinst = TipoInstitucion.objects.all()
        tiporegimen = RegimenSalud.objects.all()
        tiporegpens = RegimenPensionario.objects.all()
        tipocobsalud = CoberturaSalud.objects.all()
        tipocontr = TipoContrato.objects.all()
        tipopago = TipoPago.objects.all()
        tiposuspension = Suspension.objects.all()
        tipoexamen = TipoExamen.objects.all()
        proveedor = Proveedor.objects.all()
        listtelef = PhoneEmple.objects.all()
        listdoc = TipoDocumento.objects.all()
        proyecto = Proyecto.objects.filter(status='AC')

        return render(request, 'rrhh/menu.html', {
            'list':emple,
            'ltipemp':tipemp,
            'lcargos':cargos,
            'ltipinst':tipoinst,
            'lproyecto':proyecto,
            'lregsalud':tiporegimen,
            'lregpens':tiporegpens,
            'ltipocontrato':tipocontr,
            'ltipopago':tipopago,
            'ltiposuspension':tiposuspension,
            'ltipoexamen':tipoexamen,
            'lproveedor':proveedor,
            'ltelef':listtelef,
            'ltipodocumento':listdoc,
            'lcobsalud':tipocobsalud
            })

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        try:
            context = dict()
            if request.is_ajax():
                try:
                    if 'exists' in request.POST:
                        try:
                            Employee.objects.get(
                                pk=request.POST.get('dni'))
                            context['status'] = True
                        except Employee.DoesNotExist, e:
                            context['status'] = False

                    if 'saveemployee' in request.POST:
                        fotodoc = request.POST.get('cambfoto')
                        Employee.objects.create(
                            firstname=request.POST.get('nombre'),
                            lastname=request.POST.get('apellidos'),
                            empdni_id=request.POST.get('dni'),
                            email=request.POST.get('email'),
                            sexo= True if request.POST.get('sex') == 'true' else False,
                            address=request.POST.get('direc1'),
                            address2=request.POST.get('direc2'),
                            nacionalidad= True if request.POST.get('nac') == 'true' else False,
                            estadocivil=True if request.POST.get('estcivil') == 'true' else False,
                            birth=request.POST.get('fechnac'),
                            feching=request.POST.get('feching'),
                            estadoplanilla= True if request.POST.get('planilla') == 'true' else False,
                            charge_id=request.POST.get('cargo'),
                            archivo=request.FILES['cambficha'],
                            foto = '' if fotodoc == "" else request.FILES['cambfoto'],
                            discapacidad=True if request.POST.get('discap') == 'true' else False,
                            nacdpt=request.POST.get('dptnac'),
                            nacprov=request.POST.get('provnac'),
                            nacdist=request.POST.get('distnac'),
                            tallazap=request.POST.get('tzapat'),
                            tallapolo=request.POST.get('tpolo'),
                            cargopostula=request.POST.get('postula'),
                            distrito=request.POST.get('distr'),
                            tipoemple_id=request.POST.get('tipotrab'))
                        context['status'] = True


                    if 'editemployee' in request.POST:
                        archivo = request.POST.get('archiv')
                        archivofoto = request.POST.get('archivfoto')
                        empledni = request.POST.get('dni')
                        filename = '/storage/examedic/%s/fichaentrada/%s' % (empledni, archivo)
                        filenamefoto = '/storage/examedic/%s/foto/%s' % (empledni, archivofoto)
                        ficha = request.POST.get('cambficha')
                        fotoe = request.POST.get('cambfoto')

                        emple = Employee.objects.get(
                            empdni_id=request.POST.get('dni'))
                        emple.firstname = request.POST.get('nombre')
                        emple.lastname = request.POST.get('apellidos')
                        emple.email = request.POST.get('email')
                        emple.sexo = True if request.POST['sex'] == 'true' else False
                        emple.address = request.POST.get('direc1')
                        emple.address2 = request.POST['direc2']
                        emple.nacionalidad = True if request.POST['nac'] == 'true' else False
                        emple.estadocivil = True if request.POST['estcivil'] == 'true' else False
                        emple.birth = request.POST['fechnac']
                        emple.feching = request.POST['feching']
                        emple.estadoplanilla = True if request.POST['planilla'] == 'true' else False
                        emple.charge_id = request.POST['cargo']
                        if ficha != '':
                            deleteFile(filename, True)
                            if 'cambficha' in request.FILES:
                                emple.archivo = request.FILES['cambficha']
                        if fotoe != '':
                            deleteFile(filenamefoto, True)
                            if 'cambfoto' in request.FILES:
                                emple.foto = request.FILES['cambfoto']
                        emple.discapacidad = True if request.POST['discap'] == 'true' else False
                        emple.tipoemple_id = request.POST['tipotrab']
                        emple.nacdpt = request.POST.get('dptnac')
                        emple.nacprov = request.POST.get('provnac')
                        emple.nacdist = request.POST.get('distnac')
                        emple.tallazap = request.POST.get('tzapat')
                        emple.tallapolo = request.POST.get('tpolo')
                        emple.cargopostula = request.POST.get('postula')
                        emple.distrito = request.POST.get('distr')
                        emple.save()
                        context['status'] = True

                except Exception, e:
                    context['raise'] = str(e)
                    context['status'] = False
                return self.render_to_json_response(context)
        except Exception, e:
            print e.__str__()


class Proyectos(JSONResponseMixin, TemplateView):

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        if request.is_ajax():
            try:
                if 'lempxobra' in request.GET:
                    listaobra = []
                    conta = 1
                    for x in EmpleCampo.objects.filter(proyecto_id=request.GET.get('dat')):
                        listaobra.append({
                            'dni': x.empdni_id,
                            'ape': x.empdni.lastname,
                            'nomb': x.empdni.firstname,
                            'conta': conta,
                            'finiproyecto': x.proyecto.comienzo,
                            'ffinproyecto': x.proyecto.fin,
                            'carnet': 'SI' if x.carnetmag == True else 'NO',
                            'fotocheck': 'SI' if x.fotocheck == True else 'NO',
                            'proyectocod' : x.proyecto_id,
                            'finiproy':x.fechinicio})
                        conta = conta+1
                    if len(listaobra) == 0:
                        context['lobras'] = False
                    else:
                        context['lobras'] = listaobra
                    context['status'] = True

                if 'lempxinduccion' in request.GET:
                    lisinducciones = []
                    conta = 1
                    for x in Induccion.objects.filter(
                            emplecampo__proyecto_id=request.GET.get('datindxobr')):
                        lisinducciones.append({
                            'finicio':x.fechinicio,
                            'fcad':x.fechcaduca,
                            'cargo':x.emplecampo.empdni.charge.cargos,
                            'apellidos':x.emplecampo.empdni.lastname,
                            'nombres': x.emplecampo.empdni.firstname,
                            'dni':x.emplecampo.empdni_id,
                            'conta':conta,
                            'estad':x.estado})
                        conta = conta+1
                    if len(lisinducciones) == 0:
                        context['lindxobra'] = False
                    else:
                        context['lindxobra'] = lisinducciones
                    context['status'] = True


            except ObjectDoesNotExist, e:
                context['raise'] = str(e)
                context['status'] = False
            return self.render_to_json_response(context)

        proyecto = Proyecto.objects.filter(status='AC')
        return render(request, 'rrhh/proyecto.html', {'lproyecto': proyecto})


class Examenes(JSONResponseMixin, TemplateView):

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        if request.is_ajax():
            try:
                if 'lempexamen' in request.GET:
                    lisexamenes = []
                    count = 1
                    for x in ExamenEmple.objects.filter(
                            tipoexamen_id=request.GET.get('codlisexa')).order_by('-fechcaduca'):
                        context['tipoexa'] = x.tipoexamen.descripcion
                        lisexamenes.append({
                            'ape': x.empdni.lastname,
                            'nomb': x.empdni.firstname,
                            'apt': x.aptitud,
                            'lugar': x.lugar.razonsocial,
                            'fini': x.fechinicio,
                            'count': count,
                            'fcad': x.fechcaduca})
                        count = count + 1
                    if len(lisexamenes) == 0:
                        context['lempexa'] = False
                    else:
                        context['lempexa'] = lisexamenes
                    context['status'] = True

                if 'lempdoc' in request.GET:
                    lisdoc = []
                    count = 1
                    for x in DocumentoEmple.objects.filter(
                            documento_id=request.GET.get('codlisdoc')).order_by('-fechcaduca'):
                        context['tipodoc'] = x.documento.descripcion
                        lisdoc.append({
                            'apellid':x.empdni.lastname,
                            'nombr':x.empdni.firstname,
                            'cond':x.condicion,
                            'counter':count,
                            'fcaduci':x.fechcaduca})
                        count = count + 1
                    if len(lisdoc) == 0:
                        context['lempdoc'] = False
                    else:
                        context['lempdoc'] = lisdoc
                    context['status'] = True

            except ObjectDoesNotExist, e:
                context['raise'] = str(e)
                context['status'] = False
            return self.render_to_json_response(context)

        examenes = TipoExamen.objects.filter(flag='True')
        docs = TipoDocumento.objects.filter(flag='True')
        return render(request, 'rrhh/examenes.html', {'lexa': examenes, 'ldoc': docs})


class SegSocial(JSONResponseMixin, TemplateView):
    """docstring for SegSocial"""

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        if request.is_ajax():
            try:
                if 'listsegsocial' in request.GET:
                    listaregsalud = []
                    for x in detSaludEmple.objects.filter(
                            empdni_id=request.GET.get('idsegsocial')).order_by('-fechfin'):
                        listaregsalud.append({
                            'id': x.id,
                            'regimen_id' : x.regimen_id,
                            'regimen' : x.regimen.regimen,
                            'empdni_id' : x.empdni_id,
                            'nameregsal' : x.empdni.lastname,
                            'finicioregsal': x.fechinicio,
                            'ffinregsal': x.fechfin,
                            'entidad':x.entidad})
                        context['lregsalud'] = listaregsalud

                    listaregpens = []
                    for x in detPensEmple.objects.filter(
                            empdni_id=request.GET.get('idsegsocial')).order_by('-fechfin'):
                        listaregpens.append({
                            'id': x.id,
                            'regimenpens_id': x.regimenpens_id,
                            'regimenpens': x.regimenpens.regimen,
                            'empdni_id': x.empdni_id,
                            'nameregpens': x.empdni.lastname,
                            'finiregpens': x.fechinicio,
                            'ffinregpens': x.fechfin,
                            'cuspp': x.cuspp,
                            'sctr': x.sctr,
                            'vistasctr': 'SI' if x.sctr == True else 'NO'})
                        context['lregpens'] = listaregpens

                    listacobsalud = []
                    for x in detCobSaludEmple.objects.filter(
                            empdni_id=request.GET.get('idsegsocial')).order_by('-fechfin'):
                        listacobsalud.append({
                            'id': x.id,
                            'cobertura_id': x.cobertura_id,
                            'cobertura': x.cobertura.cobertura,
                            'empdni_id': x.empdni_id,
                            'namecobsal': x.empdni.lastname,
                            'finicobsal': x.fechinicio,
                            'ffincobsal':x.fechfin})
                        context['lcobsalud'] = listacobsalud
                    context['status'] = True

                if 'listfamilia' in request.GET:
                    listafamilia = []
                    for x in FamiliaEmple.objects.filter(
                            empdni_id=request.GET.get('employeeidfam')):
                        listafamilia.append({
                            'id': x.id,
                            'empdni_id' : x.empdni_id,
                            'namefam': x.empdni.lastname,
                            'nombres' : x.nombres,
                            'parentesco' : x.parentesco,
                            'edad' : x.edad,
                            'fnac':x.fnac})
                        context['lfamilia'] = listafamilia
                    context['status'] = True

                if 'listamedic' in request.GET:
                    listamed = []
                    for x in DatoMedico.objects.filter(
                            empdni_id=request.GET.get('employeeidmedic')):
                        listamed.append({
                            'id': x.id,
                            'empdni_id': x.empdni_id,
                            'namemedic':x.empdni.lastname,
                            'tipo': x.tipo,
                            'descripcion': x.descripcion,
                            'tiempo': x.tiempo,
                            'comentario':x.comentario})
                        context['lmedic'] = listamed
                    context['status'] = True

                if 'listaemerg' in request.GET:
                    listemerg = []
                    for x in CasoEmergencia.objects.filter(
                            empdni_id=request.GET.get('employeeidemerg')):
                        listemerg.append({
                            'id': x.id,
                            'empdni_id': x.empdni_id,
                            'nameemerg':x.empdni.lastname,
                            'nombres': x.nombres,
                            'direccion': x.direccion,
                            'telefono': x.telefono,
                            'parentesco':x.parentesco})
                        context['lemerg'] = listemerg
                    context['status'] = True
                if 'listafamicr' in request.GET:
                    listfamicr = []
                    for x in FamiliaIcr.objects.filter(
                            empdni_id=request.GET.get('employeeidfamicr')):
                        listfamicr.append({
                            'id': x.id,
                            'empdni_id': x.empdni_id,
                            'namefamicr':x.empdni.lastname,
                            'nombres': x.nombres,
                            'parent': x.parentesco,
                            'areatrab':x.areatrabajo})
                        context['lfamicr'] = listfamicr
                    context['status'] = True

                # if 'lempxobra' in request.GET:
                #     listaobra = []
                #     for x in EmpleCampo.objects.filter(proyecto_id=request.GET.get('dat')):
                #         listaobra.append(
                #             {'dni': x.empdni_id,
                #             'proyectocod' : x.proyecto_id,
                #             'finiproy':x.fechinicio})
                #         context['lobras']= listaobra
                #     context['status'] = True

            except ObjectDoesNotExist, e:
                context['raise'] = str(e)
                context['status'] = False
            return self.render_to_json_response(context)

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        try:
            context = dict()
            if request.is_ajax():
                try:
                    if 'savesegsocial' in request.POST:
                        detPensEmple.objects.create(
                            regimenpens_id=request.POST.get('regpens'),
                            empdni_id=request.POST.get('dniemple'),
                            fechinicio=request.POST.get('regpensfechini'),
                            fechfin= None if request.POST.get('regpensfechfin') == '' else request.POST.get('regpensfechfin'),
                            cuspp=request.POST.get('regpenscuspp'),
                            sctr=True if request.POST.get('sctr') == 'true' else False)

                        detCobSaludEmple.objects.create(
                            cobertura_id=request.POST.get('cobsalud'),
                            empdni_id=request.POST.get('dniemple'),
                            fechinicio=request.POST.get('cobsalfechini'),
                            fechfin= None if request.POST.get('cobsalfechfin') == '' else request.POST.get('cobsalfechfin'))
                        detSaludEmple.objects.create(
                            regimen_id=request.POST.get('regsalud'),
                            empdni_id=request.POST.get('dniemple'),
                            fechinicio=request.POST.get('regsalfechini'),
                            fechfin= None if request.POST.get('regsalfechfin') == '' else request.POST.get('regsalfechfin'),
                            entidad=request.POST.get('regsalentidad'))
                        context['status'] = True

                    if 'saveeditregsalud' in request.POST:
                        regsal = detSaludEmple.objects.get(
                            id=request.POST.get('codregsalud'),
                            empdni_id=request.POST.get('dni'))
                        regsal.regimen_id = request.POST.get('regimensalud')
                        regsal.fechinicio = request.POST.get('finiregimensalud')
                        regsal.fechfin = None if request.POST.get('ffinregimensalud') == '' else request.POST.get('ffinregimensalud')
                        regsal.entidad = request.POST.get('regsaludentidad')
                        regsal.save()
                        context['status'] = True

                    if 'delregsalud' in request.POST:
                        detSaludEmple.objects.get(
                            empdni_id=request.POST.get('dni'),
                            id=request.POST.get('pk')).delete()
                        context['status'] = True

                    if 'saveeditregpens' in request.POST:
                        regpens = detPensEmple.objects.get(
                            id=request.POST.get('codregpens'),
                            empdni_id=request.POST.get('dni')
                            )
                        regpens.regimenpens_id = request.POST.get('regimenpensio')
                        regpens.fechinicio = request.POST.get('finiregimenpens')
                        regpens.fechfin = None if request.POST.get('ffinregimenpens') == '' else request.POST.get('ffinregimenpens')
                        regpens.cuspp = request.POST.get('regipenscuspp')
                        regpens.sctr = True if request.POST.get('regimenpenssctr') == 'true' else False
                        regpens.save()
                        context['status'] = True

                    if 'delregpensionario' in request.POST:
                        detPensEmple.objects.get(
                            empdni_id=request.POST.get('dni'),
                            id=request.POST.get('pk')).delete()
                        context['status'] = True

                    if 'saveeditcobsalud' in request.POST:
                        cobsal = detCobSaludEmple.objects.get(
                            id=request.POST.get('codcobsal'),
                            empdni_id=request.POST.get('dni'))
                        cobsal.cobertura_id = request.POST.get('cobertsalud')
                        cobsal.fechinicio = request.POST.get('finicobertsalud')
                        cobsal.fechfin = None if request.POST.get('ffincobersalud') == "" else request.POST.get('ffincobersalud')
                        cobsal.save()
                        context['status'] = True

                    if 'delcobertsalud' in request.POST:
                        detCobSaludEmple.objects.get(
                            empdni_id=request.POST.get('dni'),
                            id=request.POST.get('pk')).delete()
                        context['status'] = True

                    if 'savefamilia' in request.POST:
                        FamiliaEmple.objects.create(
                            empdni_id=request.POST.get('dni'),
                            nombres=request.POST.get('nombresfam'),
                            parentesco=request.POST.get('parentfam'),
                            edad=request.POST.get('edadfam'),
                            fnac=request.POST.get('fnacfam'))
                        context['status'] = True

                    if 'saveeditfamilia' in request.POST:
                        fam = FamiliaEmple.objects.get(
                            id=request.POST.get('codfamilia'),
                            empdni_id=request.POST.get('dni'))
                        fam.nombres = request.POST.get('nombresfam')
                        fam.parentesco = request.POST.get('parentfam')
                        fam.edad = request.POST.get('edadfam')
                        fam.fnac = request.POST.get('fnacfam')
                        fam.save()
                        context['status'] = True

                    if 'delfam' in request.POST:
                        FamiliaEmple.objects.get(
                            empdni_id=request.POST.get('dni'),
                            id=request.POST.get('pk')).delete()
                        context['status'] = True

                    if 'savemedic' in request.POST:
                        DatoMedico.objects.create(
                            empdni_id=request.POST.get('dni'),
                            tipo=request.POST.get('tipomedic'),
                            descripcion=request.POST.get('descmedic'),
                            tiempo=request.POST.get('timemedic'),
                            comentario=request.POST.get('commedic'))
                        context['status'] = True

                    if 'saveeditmedic' in request.POST:
                        med = DatoMedico.objects.get(
                            id=request.POST.get('codme'),
                            empdni_id=request.POST.get('dni'))
                        med.tipo = request.POST.get('tipomedic')
                        med.descripcion = request.POST.get('descmedic')
                        med.tiempo = request.POST.get('timemedic')
                        med.comentario = request.POST.get('commedic')
                        med.save()
                        context['status'] = True

                    if 'delmedic' in request.POST:
                        DatoMedico.objects.get(
                            empdni_id=request.POST.get('dni'),
                            id=request.POST.get('pk')).delete()
                        context['status'] = True

                    if 'saveemerg' in request.POST:
                        CasoEmergencia.objects.create(
                            empdni_id=request.POST.get('dni'),
                            nombres=request.POST.get('emergnombres'),
                            direccion=request.POST.get('emergdirec'),
                            telefono=request.POST.get('emergtel'),
                            parentesco=request.POST.get('emergparent'))
                        context['status'] = True

                    if 'saveeditemerg' in request.POST:
                        emerg = CasoEmergencia.objects.get(
                            id=request.POST.get('codemergencia'),
                            empdni_id=request.POST.get('dni'))
                        emerg.nombres = request.POST.get('emergnombres')
                        emerg.direccion = request.POST.get('emergdirec')
                        emerg.telefono = request.POST.get('emergtel')
                        emerg.parentesco = request.POST.get('emergparent')
                        emerg.save()
                        context['status'] = True

                    if 'delemerg' in request.POST:
                        CasoEmergencia.objects.get(
                            empdni_id=request.POST.get('dni'),
                            id=request.POST.get('pk')).delete()
                        context['status'] = True

                    if 'savefamicr' in request.POST:
                        FamiliaIcr.objects.create(
                            empdni_id=request.POST.get('dni'),
                            nombres=request.POST.get('famicrnombres'),
                            parentesco=request.POST.get('famicrparent'),
                            areatrabajo=request.POST.get('famicrareatrab'))
                        context['status'] = True

                    if 'saveeditfamicr' in request.POST:
                        famicr = FamiliaIcr.objects.get(
                            id=request.POST.get('codfamiliaicr'),
                            empdni_id=request.POST.get('dni'))
                        famicr.nombres = request.POST.get('famicrnombres')
                        famicr.parentesco = request.POST.get('famicrparent')
                        famicr.areatrabajo = request.POST.get('famicrareatrab')
                        famicr.save()
                        context['status'] = True

                    if 'delfamiliaicr' in request.POST:
                        FamiliaIcr.objects.get(
                            empdni_id=request.POST.get('dni'),
                            id=request.POST.get('pk')).delete()
                        context['status'] = True

                    if 'savemotren' in request.POST:
                        docren = request.POST.get('docrenuncia')
                        RenunciaEmple.objects.create(
                            empdni_id=request.POST.get('dni'),
                            motivo=request.POST.get('motrenun'),
                            ultimodiatrab=request.POST.get('ultdiatrab'),
                            archivo= '' if docren == "" else request.FILES['docrenuncia'])
                        context['status'] = True

                    if 'cambestemple' in request.POST:
                        cambest = Employee.objects.get(
                            empdni_id=request.POST.get('dni'))
                        cambest.flag = False
                        cambest.save()
                        context['status'] = True

                    if 'activaremple' in request.POST:
                        actemp = Employee.objects.get(
                            empdni_id=request.POST.get('dni'))
                        actemp.flag = True
                        actemp.save()
                        context['status'] = True

                except Exception, e:
                    context['raise'] = str(e)
                    context['status'] = False
                return self.render_to_json_response(context)
        except Exception, e:
            print e.__str__()


class Phone(JSONResponseMixin, TemplateView):
    """docstring for Menu"""
    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        if request.is_ajax():
            try:
                if 'listphone' in request.GET:
                    listatelef = []
                    for x in PhoneEmple.objects.filter(
                            empdni_id=request.GET.get('employeeidphone')):
                        listatelef.append({
                            'id': x.id,
                            'empdni_id' : x.empdni_id,
                            'nametel': x.empdni.lastname,
                            'phone': x.phone,
                            'descripcion': x.descripcion})
                        context['listaphone'] = listatelef

                    context['status'] = True
            except ObjectDoesNotExist, e:
                context['raise'] = str(e)
                context['status'] = False
            return self.render_to_json_response(context)

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        try:
            context = dict()
            if request.is_ajax():
                try:
                    if 'exists' in request.POST:
                        try:
                            PhoneEmple.objects.get(
                                phone=request.POST.get('telef'))
                            context['status'] = True
                        except PhoneEmple.DoesNotExist, e:
                            context['status'] = False
                    if 'savephone' in request.POST:
                        PhoneEmple.objects.create(
                            empdni_id=request.POST.get('dni'),
                            phone=request.POST.get('telef'),
                            descripcion=request.POST.get('descripcion')
                            )
                        context['status'] = True

                    if 'edittelefono' in request.POST:
                        pe = PhoneEmple.objects.get(
                            id=request.POST.get('codtelef'),
                            empdni_id=request.POST.get('dni')
                            )
                        pe.phone = request.POST.get('telef')
                        pe.descripcion = request.POST.get('descripcion')
                        pe.save()
                        context['status'] = True

                    if 'delphoneemple' in request.POST:
                        PhoneEmple.objects.get(
                            empdni_id=request.POST.get('dni'),
                            phone=request.POST.get('phone')).delete()
                        context['status'] = True

                except Exception, e:
                    context['raise'] = str(e)
                    context['status'] = False
                return self.render_to_json_response(context)
        except Exception, e:
            print e.__str__()


class CuentaCorriente(JSONResponseMixin, TemplateView):
    """docstring for Menu"""
    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        if request.is_ajax():
            try:
                if 'licuenta' in request.GET:
                    listacuenta = []
                    for x in CuentaEmple.objects.filter(
                            empdni_id=request.GET.get('employeeidcuenta')).order_by('remuneracion'):
                        listacuenta.append({
                            'id': x.id,
                            'empdni_id': x.empdni_id,
                            'namecuen': x.empdni.lastname,
                            'cuenta': x.cuenta,
                            'tipodepago_id': x.tipodepago_id,
                            'tipodepago': x.tipodepago.pago,
                            'remuneracion': x.remuneracion,
                            'tipocontrato_id': x.tipocontrato_id,
                            'tipocontrato': x.tipocontrato.contrato,
                            'cts': x.cts,
                            'gratificacion': x.gratificacion,
                            'costxhora': x.costxhora,
                            'estado': x.estado,
                            'setfamily': x.setfamily})
                        context['lcuenta'] = listacuenta
                    context['status'] = True
            except ObjectDoesNotExist, e:
                context['raise'] = str(e)
                context['status'] = False
            return self.render_to_json_response(context)

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        try:
            context = dict()
            if request.is_ajax():
                try:
                    if 'exists' in request.POST:
                        try:
                            CuentaEmple.objects.get(
                                cuenta=request.POST.get('numcuenta'))
                            context['status'] = True
                        except CuentaEmple.DoesNotExist, e:
                            context['status'] = False
                    if 'savecuenta' in request.POST:
                        CuentaEmple.objects.create(
                            empdni_id=request.POST.get('dni'),
                            cuenta=request.POST.get('numcuenta'),
                            tipodepago_id=request.POST.get('tipopago'),
                            estado=request.POST.get('cueactivo'),
                            remuneracion=request.POST.get('remuneracion'),
                            gratificacion=request.POST.get('gratificacion'),
                            cts=request.POST.get('cts'),
                            costxhora=request.POST.get('costoxhora'),
                            setfamily=request.POST['setfamily'],
                            tipocontrato_id=request.POST.get('tipocontrato'))
                        context['status'] = True
                    if 'editcuenta' in request.POST:
                        ce = CuentaEmple.objects.get(
                            id=request.POST.get('cod'),
                            empdni_id=request.POST.get('dni'))
                        ce.cuenta = request.POST.get('numcuenta')
                        ce.tipodepago_id = request.POST.get('tipopago')
                        # ce.estado = request.POST.get('cueactivo')
                        ce.remuneracion = request.POST.get('remuneracion')
                        ce.gratificacion = request.POST.get('gratificacion')
                        ce.cts = request.POST['cts']
                        ce.costxhora = request.POST['costoxhora']
                        ce.tipocontrato_id = request.POST['tipocontrato']
                        ce.setfamily = True if request.POST['setfamily'] == 'true' else False 
                        ce.save()
                        context['status'] = True
                    if 'cambestadocuenta' in request.POST:
                        cmbcuenta = CuentaEmple.objects.get(
                            id=request.POST.get('cmbnumcuenta'),
                            empdni_id=request.POST.get('dni')
                            )
                        cmbcuenta.estado = request.POST.get('inactivo')
                        cmbcuenta.save()
                        context['status'] = True

                    if 'delcuentaemple' in request.POST:
                        CuentaEmple.objects.get(
                            empdni_id=request.POST.get('dni'),
                            cuenta=request.POST.get('cuenta')).delete()
                        context['status'] = True

                except Exception, e:
                    context['raise'] = str(e)
                    context['status'] = False
                return self.render_to_json_response(context)
        except Exception, e:
            print e.__str__()

class EstudiosEmpleado(JSONResponseMixin, TemplateView):
    """docstring for Menu"""
    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        if request.is_ajax():
            try:
                if 'liestudios' in request.GET:
                    listaestudios = []
                    for x in detEstudioEmple.objects.filter(
                            empdni_id=request.GET.get('employeeidest')):
                        listaestudios.append({
                            'id': x.id,
                            'empdni_id': x.empdni_id,
                            'nameest': x.empdni.lastname,
                            'carrera': x.carrera,
                            'anoEgreso': x.anoEgreso,
                            'situacioneduc': x.situacioneduc,
                            'estpais': x.estpais,
                            'estudioperu' : 'SI' if x.estpais == True else 'NO',
                            'tipoinstitucion_id': x.tipoinstitucion_id,
                            'tipoinst': x.tipoinstitucion.tipo,
                            'institucion': x.institucion,
                            'fechini' : x.finicio,
                            'fechfin': x.ffin,
                            'regimen':x.regimen})
                        context['lestudios'] = listaestudios
                    context['status'] = True
                if 'listexplab' in request.GET:
                    listaexplab = []
                    for x in ExperienciaLab.objects.filter(
                            empdni_id=request.GET.get('employeeidexplab')):
                        listaexplab.append({
                            'id': x.id,
                            'empdni_id': x.empdni_id,
                            'nameexplab': x.empdni.lastname,
                            'empresa': x.empresa,
                            'cargo': x.cargo,
                            'finicio': x.finicio,
                            'ffin': x.ffin,
                            'duracion': x.duracion,
                            'motivoretiro':x.motivoretiro})
                        context['lexplab'] = listaexplab
                    context['status'] = True
            except ObjectDoesNotExist, e:
                context['raise'] = str(e)
                context['status'] = False
            return self.render_to_json_response(context)

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        try:
            context = dict()
            if request.is_ajax():
                try:
                    if 'saveestudio' in request.POST:
                        detEstudioEmple.objects.create(
                            empdni_id=request.POST.get('dni'),
                            tipoinstitucion_id=request.POST.get('tipoinst'),
                            carrera=request.POST.get('carrera'),
                            anoEgreso=request.POST.get('aegreso'),
                            situacioneduc=request.POST.get('siteduc'),
                            estpais= True if request.POST.get('estpais') == 'true' else False,
                            institucion=request.POST.get('inst'),
                            finicio=request.POST.get('fini'),
                            ffin=request.POST.get('ffin'),
                            regimen=request.POST.get('instreg'))
                        context['status'] = True
                    if 'saveexplab' in request.POST:
                        ExperienciaLab.objects.create(
                            empdni_id=request.POST.get('dni'),
                            empresa=request.POST.get('explabemp'),
                            cargo=request.POST.get('explabcargo'),
                            finicio=request.POST.get('explabini'),
                            ffin=request.POST.get('explabfin'),
                            duracion=request.POST.get('explabdur'),
                            motivoretiro=request.POST.get('explabret'))
                        context['status'] = True
                    if 'editestudio' in request.POST:
                        dest = detEstudioEmple.objects.get(
                            id=request.POST.get('codestudio'),
                            empdni_id=request.POST.get('dni'))
                        dest.carrera = request.POST.get('carrera')
                        dest.anoEgreso = request.POST.get('aegreso')
                        dest.situacioneduc = request.POST.get('siteduc')
                        dest.estpais = True if request.POST.get('estpais') == 'true' else False
                        dest.tipoinstitucion_id = request.POST.get('tipoinst')
                        dest.institucion = request.POST['inst']
                        dest.regimen = request.POST['instreg']
                        dest.finicio = request.POST.get('fini')
                        dest.ffin = request.POST.get('ffin')
                        dest.save()
                        context['status'] = True
                    if 'saveeditexplab' in request.POST:
                        explab = ExperienciaLab.objects.get(
                            id=request.POST.get('codexplab'),
                            empdni_id=request.POST.get('dni'))
                        explab.empresa = request.POST.get('explabemp')
                        explab.cargo = request.POST.get('explabcargo')
                        explab.finicio = request.POST.get('explabini')
                        explab.ffin = request.POST.get('explabfin')
                        explab.motivoretiro = request.POST.get('explabret')
                        explab.duracion = request.POST.get('explabdur')
                        explab.save()
                        context['status'] = True
                    if 'delestudioemple' in request.POST:
                        detEstudioEmple.objects.get(
                            empdni_id=request.POST.get('dni'),
                            id=request.POST.get('pk')).delete()
                        context['status'] = True
                    if 'delexplaboral' in request.POST:
                        ExperienciaLab.objects.get(
                            empdni_id=request.POST.get('dni'),
                            id=request.POST.get('pk')).delete()
                        context['status'] = True
                except Exception, e:
                    context['raise'] = str(e)
                    context['status'] = False
                return self.render_to_json_response(context)
        except Exception, e:
            print e.__str__()


class SuspensionEmple(JSONResponseMixin, TemplateView):
    """docstring for Menu"""
    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        if request.is_ajax():
            try:
                if 'lisuspension' in request.GET:
                    listasuspension = []
                    for x in detSuspension.objects.filter(
                            empdni_id=request.GET.get('employeeidsusp')).order_by('-fechfin'):
                        if x.estado == 'ACTIVO':
                            listasuspension.append({
                                'empdni_id': x.empdni_id,
                                'namesusp': x.empdni.lastname,
                                'id': x.id,
                                'suspension_id': x.suspension_id,
                                'motivo': x.suspension.motivo,
                                'fechinicio': x.fechinicio,
                                'archisusp': str(x.archivo),
                                'fechfin': x.fechfin,
                                'estado': x.estado,
                                'comentario':x.comentario})
                            context['lsuspension'] = listasuspension
                    context['status'] = True
            except ObjectDoesNotExist, e:
                context['raise'] = str(e)
                context['status'] = False
            return self.render_to_json_response(context)

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        try:
            context = dict()
            if request.is_ajax():
                try:
                    if 'savesuspension' in request.POST:
                        suspens = request.POST.get('cambsusp')
                        detSuspension.objects.create(
                            empdni_id=request.POST.get('dni'),
                            suspension_id=request.POST.get('motivo'),
                            fechinicio=request.POST.get('fechinicio'),
                            fechfin=request.POST.get('fechfinsusp'),
                            estado=request.POST.get('estadosusp'),
                            archivo= '' if suspens == "" else request.FILES['cambsusp'],
                            comentario=request.POST.get('comentario'))
                        context['status'] = True
                    if 'editsus' in request.POST:
                        archivo = request.POST.get('archi')
                        empledni = request.POST.get('dni')
                        filename = '/storage/examedic/%s/suspension/%s' % (empledni,archivo)
                        susp = request.POST.get('cambsusp')
                        dsus = detSuspension.objects.get(
                            id=request.POST.get('codsu'),
                            empdni_id=request.POST.get('dni'))
                        dsus.suspension_id = request.POST.get('motivo')
                        dsus.fechinicio = request.POST.get('fechinicio')
                        dsus.fechfin = request.POST.get('fechfinsusp')
                        dsus.comentario = request.POST.get('comentario')
                        if susp != '':
                            deleteFile(filename, True)
                            dsus.archivo = request.FILES['cambsusp']
                        dsus.save()
                        context['status'] = True
                    if 'delsuspemple' in request.POST:
                        archivo = request.POST.get('urlarchi')
                        empledni = request.POST.get('dni')
                        filename = '/storage/examedic/%s/suspension/%s' % (empledni, archivo)
                        deleteFile(filename, True)
                        detSuspension.objects.get(
                            empdni_id=request.POST.get('dni'),
                            id=request.POST.get('pk')).delete()
                        context['status'] = True
                    if 'finsusp' in request.POST:
                        fins = detSuspension.objects.get(
                            id=request.POST.get('codfinsusp'),
                            empdni_id=request.POST.get('dni'))
                        fins.estado = request.POST.get('estado')
                        fins.save()
                        context['status'] = True
                except Exception, e:
                    context['raise'] = str(e)
                    context['status'] = False
                return self.render_to_json_response(context)
        except Exception, e:
            print e.__str__()


class DocumentoEmpleado(JSONResponseMixin, TemplateView):
    """docstring for Menu"""
    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        if request.is_ajax():
            try:
                if 'listexadoc' in request.GET:
                    listaexa = []
                    for x in ExamenEmple.objects.filter(
                            empdni_id=request.GET.get('idexado')):
                        listaexa.append({
                            'codexa': x.id,
                            'empdni_id': x.empdni_id,
                            'nameexa': x.empdni.lastname,
                            'tipoexa_id': x.tipoexamen_id,
                            'tipoexa' : x.tipoexamen.descripcion,
                            'finicio': x.fechinicio,
                            'fcaduca': x.fechcaduca,
                            'apt': x.aptitud,
                            'archi': str(x.archivo),
                            'lug_id': x.lugar_id,
                            'lug': x.lugar.razonsocial,
                            'coment':x.comentario})
                        context['lexamen'] = listaexa
                    listadocumento = []
                    for y in DocumentoEmple.objects.filter(empdni_id=request.GET.get('idexado')):
                        listadocumento.append({
                            'coddocu': y.id,
                            'empdni_id' : y.empdni_id,
                            'namedoc' : y.empdni.lastname,
                            'docu': y.documento_id,
                            'namedocu': y.documento.descripcion,
                            'finicio': y.fechinicio,
                            'condic': y.condicion,
                            'archidoc': str(y.archivo),
                            'obser': y.observaciones,
                            'fcaduca': y.fechcaduca})
                        context['ldocumento'] = listadocumento
                    context['status'] = True
            except ObjectDoesNotExist, e:
                context['raise'] = str(e)
                context['status'] = False
            return self.render_to_json_response(context)

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        try:
            context = dict()
            if request.is_ajax():
                try:
                    if 'saveexamen' in request.POST:
                        examen = request.POST.get('cambexa')
                        ExamenEmple.objects.create(
                            empdni_id=request.POST.get('dni'),
                            tipoexamen_id=request.POST.get('tipexa'),
                            fechinicio=request.POST.get('finiexa'),
                            fechcaduca=None if request.POST.get('fcadexa') == '' else request.POST.get('fcadexa'),
                            aptitud=request.POST.get('aptexa'),
                            comentario=request.POST.get('comentexa'),
                            archivo= '' if examen == '' else request.FILES['cambexa'],
                            lugar_id=request.POST.get('lugexa'))
                        context['status'] = True
                    if 'savedocumento' in request.POST:
                        documento = request.POST.get('cambdocu')
                        DocumentoEmple.objects.create(
                            empdni_id=request.POST.get('dni'),
                            documento_id=request.POST.get('namedocu'),
                            fechinicio=None if request.POST.get('finidocu') == '' else request.POST.get('finidocu'),
                            fechcaduca=None if request.POST.get('fcaddocu') == '' else request.POST.get('fcaddocu'),
                            archivo='' if documento == '' else request.FILES['cambdocu'],
                            observaciones=request.POST.get('obserdocu'),
                            condicion=request.POST.get('condicdoc'))
                        context['status'] = True
                    if 'editexamen' in request.POST:
                        archivo = request.POST.get('arch')
                        empledni = request.POST.get('dni')
                        filename = '/storage/examedic/%s/examenes/%s' % (empledni, archivo)
                        exa = request.POST.get('cambexa')
                        edexa = ExamenEmple.objects.get(
                            id=request.POST.get('codexa'),
                            empdni_id=request.POST.get('dni'))
                        edexa.tipoexamen_id = request.POST.get('tipexa')
                        edexa.lugar_id = request.POST.get('lugexa')
                        edexa.fechinicio = request.POST.get('finiexa')
                        edexa.fechcaduca = None if request.POST.get('fcadexa') == "" else request.POST.get('fcadexa')
                        edexa.aptitud = request.POST.get('aptexa')
                        edexa.comentario = request.POST.get('comentexa')
                        if exa != '':
                            deleteFile(filename, True)
                            edexa.archivo = request.FILES['cambexa']
                        edexa.save()
                        context['status'] = True
                    if 'editdocum' in request.POST:
                        archivo = request.POST.get('archi')
                        empledni = request.POST.get('dni')
                        filename = '/storage/examedic/%s/documentos/%s' % (empledni, archivo)
                        doc = request.POST.get('cambdocu')
                        eddoc = DocumentoEmple.objects.get(
                            id=request.POST.get('coddocum'),
                            empdni_id=request.POST.get('dni'))
                        eddoc.documento_id = request.POST.get('namedocu')
                        eddoc.condicion = request.POST.get('condicdoc')
                        eddoc.fechinicio = request.POST.get('finidocu')
                        eddoc.fechcaduca = request.POST.get('fcaddocu')
                        eddoc.observaciones = request.POST.get('obserdocu')
                        if doc != '':
                            deleteFile(filename, True)
                            eddoc.archivo = request.FILES['cambdocu']
                        eddoc.save()
                        context['status'] = True
                    if 'delexamenemple' in request.POST:
                        archivo = request.POST.get('namearchivo')
                        empledni = request.POST.get('dni')
                        filename = '/storage/examedic/%s/examenes/%s' % (empledni, archivo)
                        deleteFile(filename, True)
                        ExamenEmple.objects.get(
                            empdni_id=request.POST.get('dni'),
                            id=request.POST.get('pk')).delete()
                        context['status'] = True
                    if 'deldocuemple' in request.POST:
                        archivo = request.POST.get('namearchdoc')
                        empledni = request.POST.get('dni')
                        filename = '/storage/examedic/%s/documentos/%s' % (empledni, archivo)
                        deleteFile(filename, True)
                        DocumentoEmple.objects.get(
                            empdni_id=request.POST.get('dni'),
                            id=request.POST.get('pk')).delete()
                        context['status'] = True
                except Exception, e:
                    context['raise'] = str(e)
                    context['status'] = False
                return self.render_to_json_response(context)
        except Exception, e:
            print e.__str__()

class ObraEmpleado(JSONResponseMixin, TemplateView):
    """docstring for Menu"""
    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        if request.is_ajax():
            try:
                if 'listinduccion' in request.GET:
                    listainduccion = []
                    for y in Induccion.objects.filter(
                            emplecampo_id=request.GET.get('employeeidinduc')):
                        listainduccion.append({
                            'codind' : y.id,
                            'emplecampo_id' : y.emplecampo_id,
                            'fechinicio': y.fechinicio,
                            'estadoinduc': y.estado,
                            'fechcaduca': y.fechcaduca,
                            'comentario':y.comentario})
                        context['linduccion'] = listainduccion
                    context['status'] = True
                if 'listobra' in request.GET:
                    listaobra = []
                    for x in EmpleCampo.objects.filter(
                            empdni_id=request.GET.get('employeeidobra')).order_by('-fechinicio'):
                        listaobra.append({
                            'codemplecam': x.emplecampo_id,
                            'empdni_id': x.empdni_id,
                            'namepro': x.empdni.lastname,
                            'proy_id': x.proyecto_id,
                            'proyecto_id': x.proyecto.nompro,
                            'fechinicio': x.fechinicio,
                            'carnet': 'SI' if x.carnetmag == True else 'NO',
                            'foto': 'SI' if x.fotocheck == True else 'NO',
                            'comentario':x.comentario})
                        context['lobra'] = listaobra
                    listaepps = []
                    li = []
                    for y in Epps.objects.filter(
                            empdni_id=request.GET.get('employeeidobra')).order_by('-fechentrega'):
                        listaepps.append({
                            'codep' : y.id,
                            'empdni_id' : y.empdni_id,
                            'nameepps': y.empdni.lastname,
                            'articulo': y.item,
                            'fechentrega': y.fechentrega,
                            'fechrecepcion': y.fechrecepcion,
                            'comentario':y.comentario})
                        context['lepps'] = listaepps
                    context['status'] = True
            except ObjectDoesNotExist, e:
                context['raise'] = str(e)
                context['status'] = False
            return self.render_to_json_response(context)

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        try:
            context = dict()
            if request.is_ajax():
                try:
                    if 'saveproyecto' in request.POST:
                        EmpleCampo.objects.create(
                            emplecampo_id = GenerateIdEmpleCampo(request.POST.get('dni'), request.POST.get('proy')),
                            empdni_id=request.POST.get('dni'),
                            proyecto_id=request.POST.get('proy'),
                            fechinicio=request.POST.get('finiproy'),
                            carnetmag= True if request.POST.get('mag') == 'true' else False,
                            fotocheck= True if request.POST.get('fot') == 'true' else False,
                            comentario=request.POST.get('comentproy'))
                        context['status'] = True
                    if 'saveinduccion' in request.POST:
                        Induccion.objects.create(
                            emplecampo_id =request.POST.get('dni'),
                            fechinicio=request.POST.get('finduc'),
                            fechcaduca=request.POST.get('fcadinduc'),
                            comentario=request.POST.get('comentinduc'))
                        context['status'] = True
                    if 'saveepps' in request.POST:
                        Epps.objects.create(
                            empdni_id =request.POST.get('dni'),
                            item=request.POST.get('arti'),
                            fechentrega=request.POST.get('fentrega'),
                            estadoepp = request.POST.get('entregado'),
                            fechrecepcion= None if request.POST.get('frecepcion') == "" else request.POST.get('frecepcion'),
                            comentario=request.POST.get('comepps'))
                        context['status'] = True
                    if 'editproy' in request.POST:
                        dempc = EmpleCampo.objects.get(
                            emplecampo_id=request.POST.get('codproy'),
                            empdni_id=request.POST.get('dni'))
                        dempc.proyecto_id = request.POST.get('proy')
                        dempc.fechinicio = request.POST.get('finiproy')
                        dempc.carnetmag = True if request.POST.get('mag') == 'true' else False
                        dempc.fotocheck = True if request.POST.get('fot') == 'true' else False
                        dempc.comentario = request.POST.get('comentproy')
                        dempc.save()
                        context['status'] = True
                    if 'editinduccion' in request.POST:
                        edind = Induccion.objects.get(
                            id=request.POST.get('codinduccion'),
                            emplecampo_id=request.POST.get('dni'))
                        edind.fechinicio = request.POST.get('finduc')
                        edind.fechcaduca = request.POST.get('fcadinduc')
                        edind.estado = request.POST.get('estinduc')
                        edind.comentario = request.POST.get('comentinduc')
                        edind.save()
                        context['status'] = True
                    if 'editepps' in request.POST:
                        eepps = Epps.objects.get(
                            id=request.POST.get('codeppss'),
                            empdni_id=request.POST.get('dni'))
                        eepps.item = request.POST.get('arti')
                        eepps.fechentrega = request.POST.get('fentrega')
                        eepps.estadoepp = request.POST.get('recepcionado')
                        eepps.fechrecepcion = None if request.POST.get('frecepcion') == "" else request.POST.get('frecepcion')
                        eepps.comentario = request.POST.get('comepps')
                        eepps.save()
                        context['status'] = True
                    if 'delproy' in request.POST:
                        EmpleCampo.objects.get(
                            empdni_id=request.POST.get('dni'),
                            emplecampo_id=request.POST.get('pk')).delete()
                        context['status'] = True
                    if 'delepps' in request.POST:
                        Epps.objects.get(
                            empdni_id=request.POST.get('dni'),
                            id=request.POST.get('pk')).delete()
                        context['status'] = True
                    if 'delinduccion' in request.POST:
                        Induccion.objects.get(
                            emplecampo_id=request.POST.get('dni'),
                            id=request.POST.get('pk')).delete()
                        context['status'] = True
                except Exception, e:
                    context['raise'] = str(e)
                    context['status'] = False
                return self.render_to_json_response(context)
        except Exception, e:
            print e.__str__()

##CRUDS##
class TipInstList(ListView):
    template_name = 'rrhh/crud/tipoinst.html'

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        if request.GET.get('menu'):
            context['menu'] = request.GET.get('menu')
        context['tipos'] = TipoInstitucion.objects.filter(flag=True)
        return render_to_response(
            self.template_name,
            context,
            context_instance=RequestContext(request))


class TipInstCreate(CreateView):
    form_class = TipoInstitucionForm
    model = TipoInstitucion
    success_url = reverse_lazy('tipoinst_list')
    template_name = 'rrhh/crud/tipoinst_form.html'

    # @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(TipInstCreate, self).dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        self.object = form.save(commit=False)
        try:
            tipo = TipoInstitucion.objects.get(tipo__icontains=self.object.tipo)
            if tipo:
                return HttpResponseRedirect(self.success_url)
        except ObjectDoesNotExist:
            self.object.tipoinst_id = GenerateIdInstituto()
            self.object.save()
            return super(TipInstCreate, self).form_valid(form)


class TipInstUpdate(UpdateView):
    form_class = TipoInstitucionForm
    model = TipoInstitucion
    slug_field = 'tipoinst_id'
    slug_url_kwarg = 'tipoinst_id'
    success_url = reverse_lazy('tipoinst_list')
    template_name = 'rrhh/crud/tipoinst_form.html'

    # @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(TipInstUpdate, self).dispatch(request, *args, **kwargs)


class TipInstDelete(DeleteView):
    model = TipoInstitucion
    slug_field = 'tipoinst_id'
    slug_url_kwarg = 'tipoinst_id'
    success_url = reverse_lazy('tipoinst_list')
    template_name = 'rrhh/crud/tipoinst_del.html'

    # @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(TipInstDelete, self).dispatch(request, *args, **kwargs)


##mant tipo empleado
class TipEmpleList(ListView):
    template_name = 'rrhh/crud/tipoemple.html'
    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        if request.GET.get('menu'):
            context['menu'] = request.GET.get('menu')
        context['tiposemple'] = TipoEmpleado.objects.filter(flag=True)
        return render_to_response(
            self.template_name,
            context,
            context_instance=RequestContext(request))


class TipEmpleCreate(CreateView):
    form_class = TipoEmpleadoForm
    model = TipoEmpleado
    success_url = reverse_lazy('tipoemple_list')
    template_name = 'rrhh/crud/tipoemple_form.html'

    # @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(TipEmpleCreate, self).dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        self.object = form.save(commit=False)
        try:
            descripcion = TipoEmpleado.objects.get(descripcion__icontains=self.object.descripcion)
            if descripcion:
                return HttpResponseRedirect(self.success_url)
        except ObjectDoesNotExist:
            self.object.tipoemple_id = GenerateIdTipEmple()
            self.object.save()
            return super(TipEmpleCreate, self).form_valid(form)


class TipEmpleUpdate(UpdateView):
    form_class = TipoEmpleadoForm
    model = TipoEmpleado
    slug_field = 'tipoemple_id'
    slug_url_kwarg = 'tipoemple_id'
    success_url = reverse_lazy('tipoemple_list')
    template_name = 'rrhh/crud/tipoemple_form.html'

    # @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(TipEmpleUpdate, self).dispatch(request, *args, **kwargs)


class TipEmpleDelete(DeleteView):
    model = TipoEmpleado
    slug_field = 'tipoemple_id'
    slug_url_kwarg = 'tipoemple_id'
    success_url = reverse_lazy('tipoemple_list')
    template_name = 'rrhh/crud/tipoemple_del.html'

    # @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(TipEmpleDelete, self).dispatch(request, *args, **kwargs)


##mant tipo COBERTURA
class TipCoberturaList(ListView):
    template_name = 'rrhh/crud/tipocobertura.html'
    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        if request.GET.get('menu'):
            context['menu'] = request.GET.get('menu')
        context['tipocobertura'] = CoberturaSalud.objects.filter(flag=True)
        return render_to_response(
            self.template_name,
            context,
            context_instance=RequestContext(request))


class TipCoberturaCreate(CreateView):
    form_class = TipoCoberturaForm
    model = CoberturaSalud
    success_url = reverse_lazy('tipocobertura_list')
    template_name = 'rrhh/crud/tipocobertura_form.html'

    # @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(TipCoberturaCreate, self).dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        self.object = form.save(commit=False)
        try:
            cobertura = CoberturaSalud.objects.get(cobertura__icontains=self.object.cobertura)
            if cobertura:
                return HttpResponseRedirect(self.success_url)
        except ObjectDoesNotExist:
            self.object.coberturasalud_id = GenerateIdCobertura()
            self.object.save()
            return super(TipCoberturaCreate, self).form_valid(form)


class TipCoberturaUpdate(UpdateView):
    form_class = TipoCoberturaForm
    model = CoberturaSalud
    slug_field = 'coberturasalud_id'
    slug_url_kwarg = 'coberturasalud_id'
    success_url = reverse_lazy('tipocobertura_list')
    template_name = 'rrhh/crud/tipocobertura_form.html'

    # @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(TipCoberturaUpdate, self).dispatch(request, *args, **kwargs)


class TipCoberturaDelete(DeleteView):
    model = CoberturaSalud
    slug_field = 'coberturasalud_id'
    slug_url_kwarg = 'coberturasalud_id'
    success_url = reverse_lazy('tipocobertura_list')
    template_name = 'rrhh/crud/tipocobertura_del.html'

    # @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(TipCoberturaDelete, self).dispatch(request, *args, **kwargs)


##mant tipo REGIMEN
class TipRegimenList(ListView):
    template_name = 'rrhh/crud/tiporegimen.html'
    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        if request.GET.get('menu'):
            context['menu'] = request.GET.get('menu')
        context['tiposregimen'] = RegimenSalud.objects.filter(flag=True)
        return render_to_response(
            self.template_name,
            context,
            context_instance=RequestContext(request))


class TipRegimenCreate(CreateView):
    form_class = TipoRegimenForm
    model = RegimenSalud
    success_url = reverse_lazy('tiporegimen_list')
    template_name = 'rrhh/crud/tiporegimen_form.html'

    # @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(TipRegimenCreate, self).dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        self.object = form.save(commit=False)
        try:
            regimen = RegimenSalud.objects.get(regimen__icontains=self.object.regimen)
            if regimen:
                return HttpResponseRedirect(self.success_url)
        except ObjectDoesNotExist:
            self.object.regimensalud_id = GenerateIdRegimenSalud()
            self.object.save()
            return super(TipRegimenCreate, self).form_valid(form)


class TipRegimenUpdate(UpdateView):
    form_class = TipoRegimenForm
    model = RegimenSalud
    slug_field = 'regimensalud_id'
    slug_url_kwarg = 'regimensalud_id'
    success_url = reverse_lazy('tiporegimen_list')
    template_name = 'rrhh/crud/tiporegimen_form.html'

    # @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(TipRegimenUpdate, self).dispatch(request, *args, **kwargs)


class TipRegimenDelete(DeleteView):
    model = RegimenSalud
    slug_field = 'regimensalud_id'
    slug_url_kwarg = 'regimensalud_id'
    success_url = reverse_lazy('tiporegimen_list')
    template_name = 'rrhh/crud/tiporegimen_del.html'

    # @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(TipRegimenDelete, self).dispatch(request, *args, **kwargs)


##mant tipo EXAMEN
class TipExamenList(ListView):
    template_name = 'rrhh/crud/tipoexamen.html'
    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        if request.GET.get('menu'):
            context['menu'] = request.GET.get('menu')
        context['tipoexamen'] = TipoExamen.objects.filter(flag=True)
        return render_to_response(
            self.template_name,
            context,
            context_instance=RequestContext(request))


class TipExamenCreate(CreateView):
    form_class = TipoExamenForm
    model = TipoExamen
    success_url = reverse_lazy('tipoexamen_list')
    template_name = 'rrhh/crud/tipoexamen_form.html'

    # @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(TipExamenCreate, self).dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        self.object = form.save(commit=False)
        try:
            descripcion = TipoExamen.objects.get(descripcion__icontains=self.object.descripcion)
            if descripcion:
                return HttpResponseRedirect(self.success_url)
        except ObjectDoesNotExist:
            self.object.tipoexamen_id = GenerateIdExamen()
            self.object.save()
            return super(TipExamenCreate, self).form_valid(form)


class TipExamenUpdate(UpdateView):
    form_class = TipoExamenForm
    model = TipoExamen
    slug_field = 'tipoexamen_id'
    slug_url_kwarg = 'tipoexamen_id'
    success_url = reverse_lazy('tipoexamen_list')
    template_name = 'rrhh/crud/tipoexamen_form.html'

    # @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(TipExamenUpdate, self).dispatch(request, *args, **kwargs)


class TipExamenDelete(DeleteView):
    model = TipoExamen
    slug_field = 'tipoexamen_id'
    slug_url_kwarg = 'tipoexamen_id'
    success_url = reverse_lazy('tipoexamen_list')
    template_name = 'rrhh/crud/tipoexamen_del.html'

    # @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(TipExamenDelete, self).dispatch(request, *args, **kwargs)


##mant tipo REGIMEN PENSIONARIO
class TipRegimenPensList(ListView):
    template_name = 'rrhh/crud/tiporegimenpens.html'
    # @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        if request.GET.get('menu'):
            context['menu'] = request.GET.get('menu')
        context['tiporegimenpens'] = RegimenPensionario.objects.filter(flag=True)
        return render_to_response(
            self.template_name,
            context,
            context_instance=RequestContext(request))


class TipRegimenPensCreate(CreateView):
    form_class = TipoRegimenPensForm
    model = RegimenPensionario
    success_url = reverse_lazy('tiporegimenpens_list')
    template_name = 'rrhh/crud/tiporegimenpens_form.html'

    # @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(TipRegimenPensCreate, self).dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        self.object = form.save(commit=False)
        try:
            regimen = RegimenPensionario.objects.get(regimen__icontains=self.object.regimen)
            if regimen:
                return HttpResponseRedirect(self.success_url)
        except ObjectDoesNotExist:
            self.object.regimenpens_id = GenerateIdRegimenPensionario()
            self.object.save()
            return super(TipRegimenPensCreate, self).form_valid(form)


class TipRegimenPensUpdate(UpdateView):
    form_class = TipoRegimenPensForm
    model = RegimenPensionario
    slug_field = 'regimenpens_id'
    slug_url_kwarg = 'regimenpens_id'
    success_url = reverse_lazy('tiporegimenpens_list')
    template_name = 'rrhh/crud/tiporegimenpens_form.html'

    # @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(TipRegimenPensUpdate, self).dispatch(request, *args, **kwargs)


class TipRegimenPensDelete(DeleteView):
    model = RegimenPensionario
    slug_field = 'regimenpens_id'
    slug_url_kwarg = 'regimenpens_id'
    success_url = reverse_lazy('tiporegimenpens_list')
    template_name = 'rrhh/crud/tiporegimenpens_del.html'

    # @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(TipRegimenPensDelete, self).dispatch(request, *args, **kwargs)


##mant tipo CONTRATO
class TipContratoList(ListView):
    template_name = 'rrhh/crud/tipocontrato.html'
    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        if request.GET.get('menu'):
            context['menu'] = request.GET.get('menu')
        context['tipocontrato'] = TipoContrato.objects.filter(flag=True)
        return render_to_response(
            self.template_name,
            context,
            context_instance=RequestContext(request))


class TipContratoCreate(CreateView):
    form_class = TipoContratoForm
    model = TipoContrato
    success_url = reverse_lazy('tipocontrato_list')
    template_name = 'rrhh/crud/tipocontrato_form.html'

    # @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(TipContratoCreate, self).dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        self.object = form.save(commit=False)
        try:
            contrato = TipoContrato.objects.get(contrato__icontains=self.object.contrato)
            if contrato:
                return HttpResponseRedirect(self.success_url)
        except ObjectDoesNotExist:
            self.object.tipocontrato_id = GenerateIdContrato()
            self.object.save()
            return super(TipContratoCreate, self).form_valid(form)


class TipContratoUpdate(UpdateView):
    form_class = TipoContratoForm
    model = TipoContrato
    slug_field = 'tipocontrato_id'
    slug_url_kwarg = 'tipocontrato_id'
    success_url = reverse_lazy('tipocontrato_list')
    template_name = 'rrhh/crud/tipocontrato_form.html'

    # @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(TipContratoUpdate, self).dispatch(request, *args, **kwargs)


class TipContratoDelete(DeleteView):
    model = TipoContrato
    slug_field = 'tipocontrato_id'
    slug_url_kwarg = 'tipocontrato_id'
    success_url = reverse_lazy('tipocontrato_list')
    template_name = 'rrhh/crud/tipocontrato_del.html'

    # @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(TipContratoDelete, self).dispatch(request, *args, **kwargs)


##mant tipo pago
class TipPagoList(ListView):
    template_name = 'rrhh/crud/tipopago.html'
    # @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        if request.GET.get('menu'):
            context['menu'] = request.GET.get('menu')
        context['tipopago'] = TipoPago.objects.filter(flag=True)
        return render_to_response(
            self.template_name,
            context,
            context_instance=RequestContext(request))


class TipPagoCreate(CreateView):
    form_class = TipoPagoForm
    model = TipoPago
    success_url = reverse_lazy('tipopago_list')
    template_name = 'rrhh/crud/tipopago_form.html'

    # @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(TipPagoCreate, self).dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        self.object = form.save(commit=False)
        try:
            pago = TipoPago.objects.get(pago__icontains=self.object.pago)
            if pago:
                return HttpResponseRedirect(self.success_url)
        except ObjectDoesNotExist:
            self.object.tipopago_id = GenerateIdPago()
            self.object.save()
            return super(TipPagoCreate, self).form_valid(form)


class TipPagoUpdate(UpdateView):
    form_class = TipoPagoForm
    model = TipoPago
    slug_field = 'tipopago_id'
    slug_url_kwarg = 'tipopago_id'
    success_url = reverse_lazy('tipopago_list')
    template_name = 'rrhh/crud/tipopago_form.html'

    # @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(TipPagoUpdate, self).dispatch(request, *args, **kwargs)


class TipPagoDelete(DeleteView):
    model = TipoPago
    slug_field = 'tipopago_id'
    slug_url_kwarg = 'tipopago_id'
    success_url = reverse_lazy('tipopago_list')
    template_name = 'rrhh/crud/tipopago_del.html'

    # @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(TipPagoDelete, self).dispatch(request, *args, **kwargs)


##mant tipo SUSPENSION
class TipSuspensionList(ListView):
    template_name = 'rrhh/crud/tiposuspension.html'
    # @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        if request.GET.get('menu'):
            context['menu'] = request.GET.get('menu')
        context['tiposuspension'] = Suspension.objects.filter(flag=True)
        return render_to_response(
            self.template_name,
            context,
            context_instance=RequestContext(request))


class TipSuspensionCreate(CreateView):
    form_class = TipoSuspensionForm
    model = Suspension
    success_url = reverse_lazy('tiposuspension_list')
    template_name = 'rrhh/crud/tiposuspension_form.html'

    # @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(TipSuspensionCreate, self).dispatch(request, *args, **kwargs)


class TipSuspensionUpdate(UpdateView):
    form_class = TipoSuspensionForm
    model = Suspension
    slug_field = 'suspension_id'
    slug_url_kwarg = 'suspension_id'
    success_url = reverse_lazy('tiposuspension_list')
    template_name = 'rrhh/crud/tiposuspension_form.html'

    # @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(TipSuspensionUpdate, self).dispatch(request, *args, **kwargs)


class TipSuspensionDelete(DeleteView):
    model = Suspension
    slug_field = 'suspension_id'
    slug_url_kwarg = 'suspension_id'
    success_url = reverse_lazy('tiposuspension_list')
    template_name = 'rrhh/crud/tiposuspension_del.html'

    # @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(TipSuspensionDelete, self).dispatch(request, *args, **kwargs)


##mant RUBRO PRO
class RubroProList(ListView):
    template_name = 'rrhh/crud/rubropro.html'
    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        if request.GET.get('menu'):
            context['menu'] = request.GET.get('menu')
        context['rubrolist'] = Rubro.objects.filter(flag=True)
        return render_to_response(
            self.template_name,
            context,
            context_instance=RequestContext(request))


class RubroProCreate(CreateView):
    form_class = RubroProForm
    model = Rubro
    success_url = reverse_lazy('rubropro_list')
    template_name = 'rrhh/crud/rubropro_form.html'

    # @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(RubroProCreate, self).dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        self.object = form.save(commit=False)
        try:
            rubro = Rubro.objects.get(rubro__icontains=self.object.rubro)
            if rubro:
                return HttpResponseRedirect(self.success_url)
        except ObjectDoesNotExist:
            self.object.rubro_id = GenerateIdRubro()
            self.object.save()
            return super(RubroProCreate, self).form_valid(form)


class RubroProUpdate(UpdateView):
    form_class = RubroProForm
    model = Rubro
    slug_field = 'rubro_id'
    slug_url_kwarg = 'rubro_id'
    success_url = reverse_lazy('rubropro_list')
    template_name = 'rrhh/crud/rubropro_form.html'

    # @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(RubroProUpdate, self).dispatch(request, *args, **kwargs)


class RubroProDelete(DeleteView):
    model = Rubro
    slug_field = 'rubro_id'
    slug_url_kwarg = 'rubro_id'
    success_url = reverse_lazy('rubropro_list')
    template_name = 'rrhh/crud/rubropro_del.html'

    # @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(RubroProDelete, self).dispatch(request, *args, **kwargs)


##mant TIPO DOCUMENTO
class TipDocList(ListView):
    template_name = 'rrhh/crud/tipodoc.html'
    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        context = dict()
        if request.GET.get('menu'):
            context['menu'] = request.GET.get('menu')
        context['doclist'] = TipoDocumento.objects.filter(flag=True)
        return render_to_response(
            self.template_name,
            context,
            context_instance=RequestContext(request))


class TipDocCreate(CreateView):
    form_class = TipDocForm
    model = TipoDocumento
    success_url = reverse_lazy('tipdoc_list')
    template_name = 'rrhh/crud/tipodoc_form.html'

    # @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(TipDocCreate, self).dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        self.object = form.save(commit=False)
        try:
            descripcion = TipoDocumento.objects.get(descripcion__icontains=self.object.descripcion)
            if descripcion:
                return HttpResponseRedirect(self.success_url)
        except ObjectDoesNotExist:
            self.object.tipodoc_id = GenerateIdDoc()
            self.object.save()
            return super(TipDocCreate, self).form_valid(form)


class TipDocUpdate(UpdateView):
    form_class = TipDocForm
    model = TipoDocumento
    slug_field = 'tipodoc_id'
    slug_url_kwarg = 'tipodoc_id'
    success_url = reverse_lazy('tipdoc_list')
    template_name = 'rrhh/crud/tipodoc_form.html'

    # @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(TipDocUpdate, self).dispatch(request, *args, **kwargs)


class TipDocDelete(DeleteView):
    model = TipoDocumento
    slug_field = 'tipodoc_id'
    slug_url_kwarg = 'tipodoc_id'
    success_url = reverse_lazy('tipdoc_list')
    template_name = 'rrhh/crud/tipodoc_del.html'

    # @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        return super(TipDocDelete, self).dispatch(request, *args, **kwargs)










class TypeEmployeeView(JSONResponseMixin, TemplateView):

    template_name = 'rrhh/typeemployee.html'

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        try:
            if request.is_ajax():
                try:
                    if 'ltypes' in request.GET:
                        kwargs['types'] = json.loads(
                            serializers.serialize(
                                'json',
                                TypesEmployee.objects.filter(flag=True).order_by('register')))
                        kwargs['status'] = True
                except ObjectDoesNotExist as oex:
                    kwargs['raise'] = str(oex)
                    kwargs['status'] = False
                return self.render_to_json_response(kwargs)
            return render(request, self.template_name, kwargs)
        except TemplateDoesNotExist as ext:
            raise Http404(ext)

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        if request.is_ajax():
            try:
                if 'new' in request.POST:
                    key = TypesEmployeeKeys()
                    TypesEmployee.objects.create(
                        types_id=key,
                        description=request.POST['desc'].upper(),
                        starthour=request.POST['starthour'],
                        outsaturday=request.POST['outsaturday'])
                    kwargs['status'] = True
                if 'modify' in request.POST:
                    obj = TypesEmployee.objects.get(types_id=request.POST['pk'])
                    obj.description = request.POST['desc'].upper()
                    obj.starthour = request.POST['starthour']
                    obj.outsaturday = request.POST['outsaturday']
                    obj.save()
                    kwargs['status'] = True
                if 'delete' in request.POST:
                    obj = TypesEmployee.objects.get(types_id=request.POST['pk'])
                    obj.delete()
                    kwargs['status'] = True
            except ObjectDoesNotExist as oex:
                kwargs['raise'] = str(oex)
                kwargs['status'] = False
            return self.render_to_json_response(kwargs)


# class for mantenice break employee
class EmployeeBreakView(JSONResponseMixin, TemplateView):
    template_name = 'rrhh/breakoemployee.html'

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        try:
            if request.is_ajax():
                try:
                    if 'getstatus' in request.GET:
                        kwargs['lstatus'] = json.loads(serializers.serialize(
                            'json',
                            EmployeeBreak.objects.filter(flag=True)))
                        kwargs['status'] = True
                except ObjectDoesNotExist as oex:
                    kwargs['raise'] = str(oex)
                    kwargs['status'] = False
                return self.render_to_json_response(kwargs)
            return render(request, self.template_name, kwargs)
        except TemplateDoesNotExist as extv:
            raise Http404(extv)

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        if request.is_ajax():
            try:
                if 'savestatus' in request.POST:
                    obj = None
                    if 'pk' in request.POST:
                        obj = EmployeeBreak.objects.get(
                            status_id=request.POST['pk'])
                    else:
                        obj = EmployeeBreak()
                        obj.status_id = StatusAssistanceId()
                    obj.description = str(request.POST['description']).upper()
                    obj.payment = request.POST['payment'] in (True, 'true', 'True')
                    obj.save()
                    kwargs['status'] = True
                if 'delete' in request.POST:
                    ebs = EmployeeBreak.objects.get(status_id=request.POST['pk'])
                    try:
                        ebs.delete()
                    except Exception as eex:
                        ebs.flag = False
                        ebs.save()
                        kwargs['msg'] = 'El estado no se ha eliminado %s' % str(eex)
                    kwargs['status'] = True
            except ObjectDoesNotExist as oex:
                kwargs['raise'] = str(oex)
                kwargs['status'] = False
            return self.render_to_json_response(kwargs)


# this class register assistenace of employee
class AssistanceEmployee(JSONResponseMixin, TemplateView):
    """this class implement asisstance employee"""
    template_name = 'rrhh/assistance.html'

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        try:
            if request.is_ajax():
                try:
                    if 'glist' in request.GET:
                        kwargs['employee'] = json.loads(
                            serializers.serialize(
                                'json',
                                Employee.objects.filter(flag=True),
                                relations=('charge')))
                        kwargs['status'] = True
                    if 'gproject' in request.GET:
                        kwargs['projects'] = json.loads(
                            serializers.serialize(
                                'json',
                                Proyecto.objects.filter(status='AC', flag=True)))
                        kwargs['status'] = True
                    if 'gettypes' in request.GET:
                        kwargs['types'] = json.loads(
                            serializers.serialize(
                                'json',
                                TypesEmployee.objects.filter(flag=True).order_by('description')))
                        kwargs['status'] = True
                except ObjectDoesNotExist as oex:
                    kwargs['raise'] = str(oex)
                    kwargs['status'] = False
                return self.render_to_json_response(kwargs)
            return  render(request, self.template_name, kwargs)
        except TemplateDoesNotExist as ext:
            raise Http404(ext)

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        if request.is_ajax():
            try:
                if 'saveAssistance' in request.POST:
                    # print request.POST
                    proj = request.POST['project'] if 'project' in request.POST else ''
                    proj = '' if proj == 'null' else proj
                    def registerassistance():
                        """this function register assistance"""
                        obj = Assistance()
                        obj.userregister_id = request.user.get_profile().empdni_id
                        obj.employee_id = request.POST['dni']
                        obj.project_id = proj if proj != '' else None
                        obj.types_id = request.POST['type']
                        obj.assistance = request.POST['date']
                        obj.hourin = request.POST['hin']
                        obj.hourout = request.POST['hout']
                        obj.hourinbreak = request.POST['hinb']
                        obj.houroutbreak = request.POST['houtb']
                        obj.viatical = request.POST['viatical']
                        obj.tag = True
                        obj.save()
                        return 1
                    exists = None
                    if proj != '' or proj is None:
                        exists = Assistance.objects.filter(
                            assistance=request.POST['date'],
                            employee_id=request.POST['dni'], project_id=proj)
                    else:
                        exists = Assistance.objects.filter(
                            assistance=request.POST['date'], employee_id=request.POST['dni'])
                    if len(exists):
                        obj = exists.latest('register')
                        hourin = request.POST['hin']
                        # print 'OBJECT ', type(obj.hourin), obj.hourin
                        hourin = str(hourin)[:5]
                        if type(hourin) is str:
                            inh = datetime.datetime.strptime(hourin, '%H:%M').time()
                        else:
                            inh = hourin
                        print type(inh), inh
                        print type(hourin), hourin
                        if inh <= obj.hourin:
                            kwargs['raise'] = 'No se puede por que la hora ingresada '\
                                'es menor a una ya ingresada para esta fecha'
                            kwargs['status'] = False
                        else:
                            registerassistance()
                    else:
                        registerassistance()
                    if 'status' not in kwargs:
                        kwargs['status'] = True
            except ObjectDoesNotExist as oex:
                kwargs['status'] = False
                kwargs['raise'] = str(oex)
            return self.render_to_json_response(kwargs)


# mantenice settings asisstance
class EmployeeAsisstanceView(JSONResponseMixin, TemplateView):

    template_name = 'rrhh/mantenice.html'

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        try:
            if request.is_ajax():
                try:
                    if 'gettypes' in request.GET:
                        kwargs['types'] = json.loads(
                            serializers.serialize(
                                'json',
                                TypesEmployee.objects.filter()
                            )
                        )
                        kwargs['status'] = True
                    if 'getsettings' in request.GET:
                        kwargs['settings'] = json.loads(
                            serializers.serialize(
                                'json',
                                EmployeeSettings.objects.filter(flag=True)
                            )
                        )
                        kwargs['status'] = True
                except ObjectDoesNotExist as exo:
                    kwargs['raise'] = str(exo)
                    kwargs['status'] = False
                return self.render_to_json_response(kwargs)
            return render(request, self.template_name, kwargs)
        except TemplateDoesNotExist as ext:
            raise Http404(ext)

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        # print request.is_ajax()
        if request.is_ajax():
            try:
                print request.POST
                if 'savesettings' in request.POST:
                    EmployeeSettings.objects.filter(flag=True).update(flag=False)
                    # print 'before declare'
                    emp = EmployeeSettings()
                    # print 'declare'
                    # print request.POST['hextperfirst'], type(request.POST['hextperfirst'])
                    emp.hextperfirst = Decimal(
                        request.POST['hextperfirst']).quantize(Decimal('0.1'))
                    emp.hextpersecond = Decimal(
                        request.POST['hextpersecond']).quantize(Decimal('0.1'))
                    emp.ngratification = request.POST['ngratification']
                    emp.ncts = request.POST['ncts']
                    emp.pergratification = Decimal(
                        request.POST['pergratification']).quantize(Decimal('0.1'))
                    emp.codeproject = request.POST['codeproject']
                    emp.starthourextra = request.POST['starthourextra']
                    emp.starthourextratwo = request.POST['starthourextratwo']
                    emp.shxsaturday = request.POST['shxsaturday']
                    emp.shxsaturdayt = request.POST['shxsaturdayt']
                    emp.totalhour = request.POST['totalhour']
                    emp.timeround = request.POST['timeround']
                    emp.save()
                    kwargs['status'] = True
            except ObjectDoesNotExist as oex:
                kwargs['raise'] = str(oex)
                kwargs['status'] = False
            return self.render_to_json_response(kwargs)


# load file for Assistance
class LoadAssistance(JSONResponseMixin, TemplateView):
    """ upload files for read and load data for the assistance """
    template_name = 'rrhh/loadfile.html'

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        try:
            if request.is_ajax():
                try:
                    pass
                except ObjectDoesNotExist as oex:
                    kwargs['raise'] = str(oex)
                    kwargs['status'] = False
                return self.render_to_json_response(kwargs)
            return render(request, self.template_name, kwargs)
        except TemplateDoesNotExist as ext:
            raise Http404(ext)

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        """ upload file """
        if request.is_ajax():
            try:
                if 'loadfiles' in request.POST:
                    # print 'Files ', request.FILES
                    valid = True
                    counter = 0
                    # load and validate format file
                    try:
                        # upload file in the path
                        path = '\\storage\\assistance\\'
                        filename = upload(path, request.FILES['files'])
                        raises = list()
                        print filename
                        # validate file is meets with format
                        wbook = load_workbook(filename=filename, data_only=True)
                        if 'Asistencia' not in wbook.get_sheet_names():
                            kwargs['raise'] = 'La hoja no existe en el archivo'
                            valid = False
                        else:
                            wsheet = wbook['Asistencia']
                            nrow = wsheet.max_row
                            ncol = wsheet.max_column
                            print 'ROW ', nrow
                            print 'COLUMN ', ncol
                            if nrow < 9:
                                valid = False
                                kwargs['raise'] = 'El archivo no cumple con el '\
                                    'numero minimo filas para registrar'
                            if ncol < 44:
                                kwargs['raise'] = 'El archivo no cumple con el numero'\
                                    ' minimo de columnas para registrar'
                                valid = False
                            if valid:
                                param = {}
                                # get code project  for employee settings
                                sett = EmployeeSettings.objects.filter(flag=True)[0]
                                def registerassistance(parameter):
                                    """this function register assistance"""
                                    try:
                                        obj = Assistance()
                                        obj.userregister_id = request.user.get_profile().empdni_id
                                        obj.employee_id = parameter['employee']
                                        obj.project_id = parameter['project']
                                        obj.types_id = parameter['types']
                                        obj.assistance = parameter['assistance']
                                        obj.hourin = parameter['hourin']
                                        obj.hourout = parameter['hourout']
                                        obj.hourinbreak = parameter['hourinbreak']
                                        obj.houroutbreak = parameter['houroutbreak']
                                        obj.viatical = parameter['viatical']
                                        obj.tag = True
                                        obj.save()
                                        print 'OK'
                                        return 1
                                    except Exception as oex:
                                        raises.append({
                                            'employee': parameter['employee'],
                                            'assistance': parameter['assistance'],
                                            'types': parameter['types'],
                                            'raise': str(oex)})
                                        return 0
                                keys = {
                                    1: 'hourin',
                                    2: 'hourinbreak',
                                    3: 'houroutbreak',
                                    4: 'hourout',
                                    5: 'viatical',
                                    6: 'project'}
                                def validformat(oparam):
                                    """ Valid format object in and fields null """
                                    print 'inside valid format '
                                    if oparam['hourinbreak'] is None:
                                        oparam['hourinbreak'] = '00:00'
                                    else:
                                        # print type(oparam['hourinbreak'])
                                        if type(oparam['hourinbreak']) is datetime:
                                            oparam['hourinbreak'] = oparam['hourinbreak'].time()
                                    if oparam['houroutbreak'] is None:
                                        oparam['houroutbreak'] = '00:00'
                                    else:
                                        if type(oparam['houroutbreak']) is datetime:
                                            oparam['houroutbreak'] = oparam['houroutbreak'].time()
                                    if oparam['hourout'] is not None:
                                        if type(oparam['hourout']) is datetime:
                                            oparam['hourout'] = oparam['hourout'].time()
                                    if oparam['hourin'] is not None:
                                        if type(oparam['hourin']) is datetime:
                                            oparam['hourin'] = oparam['hourin'].time()
                                    if oparam['viatical'] is None:
                                        oparam['viatical'] = 0
                                    print 'end format'
                                    return oparam
                                for xraw in xrange(9, nrow+1):
                                    param = {}
                                    # ndta = 3
                                    dni = wsheet.cell(row=xraw, column=2).internal_value
                                    if dni is None:
                                        continue
                                    dni = str(dni)
                                    print dni
                                    if len(dni) != 8:
                                        continue
                                    nreg = 1
                                    xcol = 2
                                    while xcol <= 44:
                                        xcol += 1
                                        if xcol > 44:
                                            break
                                        param['employee'] = dni
                                        try:
                                            if xcol % 3 == 0:
                                                assistance = wsheet.cell(row=7, column=xcol).value
                                                if assistance is not None:
                                                    print assistance
                                                    param['assistance'] = assistance
                                                else:
                                                    if nreg == 1:
                                                        xcol += 5
                                                        param = {}
                                                        continue
                                            cval = wsheet.cell(row=xraw, column=xcol).value
                                            # print cval
                                            param[keys[nreg]] = cval
                                            if nreg == 1:
                                                if param['hourin'] is None:
                                                    xcol += 5
                                                    param = {}
                                                    continue
                                            if nreg == 6:
                                                npro = ('project' in param
                                                        and param['project'] is not None)
                                                if npro:
                                                    if len(param['project']) == 7:
                                                        param['types'] = sett.codeproject
                                                    else:
                                                        if len(param['project']) == 4:
                                                            param['types'] = param['project']
                                                            param['project'] = None
                                                        else:
                                                            param['project'] = None
                                                            param['types'] = 'TY04'
                                                else:
                                                    if param['project'] is None:
                                                        param['types'] = 'TY04'
                                                    else:
                                                        if len(param['project']) == 4:
                                                            param['types'] = param['project']
                                                            param['project'] = None
                                                        else:
                                                            param['types'] = 'TY04'
                                                            param['project'] = None
                                                if param['hourin'] is not None:
                                                    nreg = 1
                                                    param = validformat(param)
                                                    print param
                                                    counter += registerassistance(param)
                                                    param = {}
                                            else:
                                                nreg += 1
                                        except Exception as exp:
                                            print exp
                                    print '============================================='
                        kwargs['status'] = valid
                    except Exception as exio:
                        kwargs['raise'] = str(exio)
                        kwargs['status'] = False
                    wbook.save(filename)
                    # time.sleep(30)
                    deleteFile(filename)
                    kwargs['raises'] = raises
                    print counter
            except ObjectDoesNotExist as oex:
                kwargs['status'] = False
                kwargs['raise'] = str(oex)
        return self.render_to_json_response(kwargs)


# Class for view assistance loaded
class ViewAssistance(JSONResponseMixin, TemplateView):

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        try:
            if request.is_ajax():
                try:
                    if 'projects' in request.GET:
                        kwargs['projects'] = json.loads(serializers.serialize(
                            'json',
                            Proyecto.objects.filter(
                                flag=True, status='AC').order_by('-registrado')))
                        kwargs['status'] = True
                    if 'gtypes' in request.GET:
                        kwargs['types'] = json.loads(serializers.serialize(
                            'json',
                            TipoEmpleado.objects.filter(flag=True)))
                        kwargs['status'] = True
                    if 'lrtypes' in request.GET:
                        kwargs['ltypes'] = json.loads(serializers.serialize(
                            'json',
                            TypesEmployee.objects.filter(flag=True)))
                        kwargs['status'] = True
                    if 'getday' in request.GET:
                        kwargs['hours'] = json.loads(serializers.serialize(
                            'json',
                            Assistance.objects.filter(
                                employee_id=request.GET['dni'],
                                assistance=request.GET['day']).order_by('hourin'),
                            relations=('project', 'types', 'status')))
                        kwargs['status'] = True
                    if 'filterAssistance' in request.GET:
                        if 'type' in request.GET:
                            ass = BalanceAssistance.objects.filter(
                                assistance__range=(
                                    request.GET['weekstart'],
                                    request.GET['weekend']),
                                employee__tipoemple_id=request.GET['type']).order_by('assistance')
                        else:
                            ass = BalanceAssistance.objects.filter(
                                assistance__range=(
                                    request.GET['weekstart'],
                                    request.GET['weekend'])).order_by('assistance')
                        week = []
                        for x in ass:
                            count = 0
                            index = 0
                            for sw in xrange(0, len(week)):
                                if week[sw]['dni'] != x.employee_id:
                                    count += 1
                                    continue
                                else:
                                    index = sw
                                    break
                            if count == len(week):
                                week.append({
                                    'dni': x.employee_id,
                                    'name': '%s %s' % (x.employee.lastname, x.employee.firstname),
                                    'days': {x.assistance.strftime('%A'): {
                                        'day': x.assistance,
                                        'hour': x.hwork,
                                        'extra': (x.hextfirst + x.hextsecond),
                                        'delay': x.hdelay,
                                        'lack': x.hlack}},
                                    'count': 1,
                                    'twork': x.hwork,
                                    'textra': (x.hextfirst + x.hextsecond),
                                    'discount': x.discount})
                            else:
                                week[index]['days'][x.assistance.strftime('%A')] = {
                                    'day': x.assistance,
                                    'hour': x.hwork,
                                    'extra': (x.hextfirst + x.hextsecond),
                                    'delay': x.hdelay,
                                    'lack': x.hlack}
                                week[index]['count'] += 1
                                week[index]['twork'] += x.hwork
                                week[index]['textra'] += (x.hextfirst + x.hextsecond)
                                week[index]['discount'] += x.discount
                        nweek = [
                            'Monday',
                            'Tuesday',
                            'Wednesday',
                            'Thursday',
                            'Friday',
                            'Saturday',
                            'Sunday']
                        for x in week:
                            for w in nweek:
                                if w not in x['days']:
                                    x['days'][w] = {'day': None}
                                else:
                                    x['viatical'] = Assistance.objects.filter(
                                        assistance=x['days'][w]['day'],
                                        employee_id=x['dni']).aggregate(
                                            Sum('viatical'))['viatical__sum']
                        kwargs['week'] = week
                        name = {}
                        day = datetime.datetime.strptime(request.GET['weekstart'], '%Y-%m-%d')
                        traslate = {
                            'Monday': 'Lunes',
                            'Tuesday': 'Martes',
                            'Wednesday': 'Miercoles',
                            'Thursday': 'Jueves',
                            'Friday': 'Viernes',
                            'Saturday': 'Sabado',
                            'Sunday': 'Domingo'}
                        for x in xrange(0, 7):
                            name[x] = {
                                'nmo': day.strftime('%A'),
                                'nmt': traslate[day.strftime('%A')],
                                'nm': day.strftime('%m/%d'),
                                'date': day.strftime('%Y-%m-%d')}
                            day = day + datetime.timedelta(days=1)
                        conf = EmployeeSettings.objects.get(flag=True)
                        kwargs['thour'] = (conf.totalhour.hour) + ((conf.totalhour.minute)/60.0)
                        kwargs['names'] = name
                        kwargs['status'] = True
                except (ObjectDoesNotExist or Exception) as oex:
                    kwargs['raise'] = str(oex)
                    kwargs['status'] = False
                return self.render_to_json_response(kwargs)
            return render(request, 'rrhh/viewassistance.html', kwargs)
        except TemplateDoesNotExist as ext:
            return Http404(ext)

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        if request.is_ajax():
            try:
                if 'saveAssistance' in request.POST:
                    if 'modifyassistance' in request.POST:
                        asl = Assistance.objects.filter(
                            employee_id=request.POST['dni'], assistance=request.POST['date'])
                        if asl:
                            try:
                                print "ready for delete assistance"
                                asl.get(pk=request.POST['pkassistance']).delete()
                                if len(asl) == 1:
                                    BalanceAssistance.objects.get(
                                        employee_id=request.POST['dni'],
                                        assistance=request.POST['date']).delete()
                            except Exception as bex:
                                pass
                    proj = request.POST['project'] if 'project' in request.POST else ''
                    proj = '' if proj == 'null' else proj
                    def registerassistance():
                        """this function register assistance"""
                        obj = Assistance()
                        obj.userregister_id = request.user.get_profile().empdni_id
                        obj.employee_id = request.POST['dni']
                        obj.project_id = proj if proj != '' else None
                        obj.types_id = request.POST['type']
                        obj.assistance = request.POST['date']
                        obj.hourin = request.POST['hin']
                        obj.hourout = request.POST['hout']
                        obj.hourinbreak = request.POST['hinb']
                        obj.houroutbreak = request.POST['houtb']
                        obj.viatical = request.POST['viatical']
                        obj.tag = True
                        obj.save()
                        return 1
                    exists = None
                    if proj != '' or proj is None:
                        exists = Assistance.objects.filter(
                            assistance=request.POST['date'],
                            employee_id=request.POST['dni'], project_id=proj)
                    else:
                        exists = Assistance.objects.filter(
                            assistance=request.POST['date'], employee_id=request.POST['dni'])
                    if len(exists):
                        obj = exists.latest('register')
                        hourin = request.POST['hin']
                        # print 'OBJECT ', type(obj.hourin), obj.hourin
                        print type(hourin), hourin
                        hourin = str(hourin)[:5]
                        if type(hourin) is str:
                            inh = datetime.datetime.strptime(hourin, '%H:%M').time()
                            # inh = datetime.strptime(hourin, '%H:%M').time()
                        else:
                            inh = hourin
                        print type(inh), inh
                        print type(hourin), hourin
                        if inh <= obj.hourin:
                            kwargs['raise'] = 'No se puede por que la hora ingresada '\
                                'es menor a una ya ingresada para esta fecha'
                            kwargs['status'] = False
                        else:
                            registerassistance()
                    else:
                        registerassistance()
                    if 'status' not in request.POST:
                        try:
                            rb = BalanceAssistance.objects.get(
                                employee_id=request.POST['dni'],
                                assistance=request.POST['date'])
                            rb.discount = request.POST['discount']
                            rb.save()
                        except Exception as rdex:
                            pass
                    kwargs['status'] = True
                if 'deleteAssistance' in request.POST:
                    asl = Assistance.objects.filter(
                        employee_id=request.POST['dni'], assistance=request.POST['date'])
                    if asl:
                        try:
                            asl.get(pk=request.POST['pk']).delete()
                            if len(asl) == 1:
                                BalanceAssistance.objects.get(
                                    employee_id=request.POST['dni'],
                                    assistance=request.POST['date']).delete()
                            else:
                                asl = Assistance.objects.filter(
                                    employee_id=request.POST['dni'],
                                    assistance=request.POST['date'])
                                asl[0].viatical = asl[0].viatical
                                asl[0].save()
                        except Exception as bex:
                            pass
                        kwargs['status'] = True
            except ObjectDoesNotExist as oex:
                kwargs['status'] = False
                kwargs['raise'] = str(oex)
            return self.render_to_json_response(kwargs)


class ExportarAssistance(JSONResponseMixin, TemplateView):

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        try:
            if request.is_ajax():
                if 'gettypes' in request.GET:
                    kwargs['types'] = json.loads(serializers.serialize(
                        'json',
                        TipoEmpleado.objects.filter(flag=True)))
                    kwargs['status'] = True
                if 'getpayment' in request.GET:
                    kwargs['payment'] = json.loads(serializers.serialize(
                        'json',
                        TipoPago.objects.filter(flag=True)))
                    kwargs['status'] = True
                return self.render_to_json_response(kwargs)
            print request.GET
            if 'exportdata' in request.GET:
                kwargs['status'] = True
                if 'type' in request.GET:
                    ass = BalanceAssistance.objects.filter(
                        assistance__range=(
                            request.GET['din'],
                            request.GET['dout']),
                        employee__tipoemple_id=request.GET['type']).order_by('employee__lastname')
                else:
                    ass = BalanceAssistance.objects.filter(
                        assistance__range=(
                            request.GET['din'],
                            request.GET['dout'])).order_by('employee__lastname')
                week = []
                for x in ass:
                    count = 0
                    index = 0
                    for sw in xrange(0, len(week)):
                        if week[sw]['dni'] != x.employee_id:
                            count += 1
                            continue
                        else:
                            index = sw
                            break
                    if count == len(week):
                        print x.employee_id
                        repen = detPensEmple.objects.get(empdni_id=x.employee_id)
                        cuenta = CuentaEmple.objects.filter(
                            empdni_id=x.employee_id).order_by('-registro')[0]
                        week.append({
                            'dni': x.employee_id,
                            'name': '%s %s' % (x.employee.lastname, x.employee.firstname),
                            'pensionario': repen.regimenpens.regimen,
                            'percent': repen.regimenpens.percent,
                            'cuenta': cuenta.cuenta,
                            'remuneracion': cuenta.remuneracion,
                            'setfamily': cuenta.setfamily,
                            # 'days': {x.assistance.strftime('%A'): {
                            #     'day': x.assistance,
                            #     'hour': x.hwork,
                            #     'extra': (x.hextfirst + x.hextsecond),
                            #     'delay': x.hdelay,
                            #     'lack': x.hlack}},
                            'count': 1,
                            'twork': x.hwork,
                            'textra': (x.hextfirst + x.hextsecond),
                            'discount': x.discount,
                            'delay': x.hdelay,
                            'viatical': Assistance.objects.filter(
                                assistance__range=(request.GET['din'], request.GET['dout']),
                                employee_id=x.employee_id).aggregate(
                                    Sum('viatical'))['viatical__sum']})
                    else:
                        # week[index]['days'][x.assistance.strftime('%A')] = {
                        #     'day': x.assistance,
                        #     'hour': x.hwork,
                        #     'extra': (x.hextfirst + x.hextsecond),
                        #     'delay': x.hdelay,
                        #     'lack': x.hlack}
                        week[index]['count'] += 1
                        week[index]['twork'] += x.hwork
                        week[index]['textra'] += (x.hextfirst + x.hextsecond)
                        week[index]['discount'] += x.discount
                        week[index]['delay'] += x.hdelay
                # Create File ExportarAssistance
                # block start style
                def style_range(ws, cell_range, options={}):
                    """
                    :param ws:  Excel worksheet instance
                    :param range: An excel range to border (e.g. A1:F20)
                    :param border: An openpyxl border object
                    """

                    start_cell, end_cell = cell_range.split(':')
                    start_coord = coordinate_from_string(start_cell)
                    start_row = start_coord[1]
                    start_col = column_index_from_string(start_coord[0])
                    end_coord = coordinate_from_string(end_cell)
                    end_row = end_coord[1]
                    end_col = column_index_from_string(end_coord[0])
                    for row in range(start_row, end_row + 1):
                        for col_idx in range(start_col, end_col + 1):
                            col = get_column_letter(col_idx)
                            if 'fnumber' in options:
                                ws.cell('%s%s' % (col, row)).number_format = options['fnumber']
                            if 'border' in options:
                                ws.cell('%s%s' % (col, row)).border = options['border']
                            if 'bg' in options:
                                ws.cell('%s%s' % (col, row)).fill = options['bg']
                            if 'alignment' in options:
                                ws.cell('%s%s' % (col, row)).alignment = options['alignment']
                # end block
                tf = {'hour': 'HORA', 'days': 'DIAS'}
                response = HttpResponse(
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment;filename=PLANILLA-%s-%s X %s.xlsx' % (
                    request.GET['din'], request.GET['dout'], tf[request.GET['tfile']])
                wb = Workbook()
                ws = wb.active
                ws.title = 'PANILLA %s AL %s' % (request.GET['din'][5:], request.GET['dout'][5:])
                ws.sheet_properties.tabColor = '1072BA'
                # set title
                ws.cell(column=1, row=3).value = 'PLANILLA DE %s AL %s' % (
                    request.GET['din'], request.GET['dout'])
                #set Header
                ws.cell(column=2, row=4).value = 'Apellidos y Nombres'
                ws.cell(column=3, row=4).value = 'Sistema de Pensiones'
                ws.cell(column=4, row=4).value = '%'
                ws.cell(column=5, row=4).value = 'Jornal'
                ws.cell(column=6, row=4).value = 'Dominical'
                ws.cell(column=7, row=4).value = 'Asig. Familiar'
                ws.cell(column=8, row=4).value = 'Total Rem.'
                ws.cell(column=9, row=4).value = '%s Laboradas' % (tf[request.GET['tfile']])
                ws.cell(column=10, row=4).value = 'Horas Extra'
                ws.cell(column=11, row=4).value = 'Rem. Total'
                ws.cell(column=12, row=4).value = 'AFP/ONP'
                ws.cell(column=13, row=4).value = 'Otros Dsct'
                ws.cell(column=14, row=4).value = 'Total Dsct'
                ws.cell(column=15, row=4).value = 'Total a Pagar'
                # set data body
                count = 5
                total = 0
                reminv = 85
                pxhour = 8
                print week
                print len(week)
                if request.GET['tfile'] == 'hour':
                    for x in week:
                        dom = 0
                        sfamily = 0
                        dom = Decimal(x['remuneracion'] / float(
                            request.GET['payment'])).quantize(Decimal('0.01'))
                        if x['setfamily'] == True:
                            sfamily = Decimal((float(reminv)/30) * float(
                                request.GET['payment'])).quantize(Decimal('0.01'))
                        ws.cell(column=1, row=count).value = (count - 4)
                        ws.cell(column=2, row=count).value = x['name']
                        ws.cell(column=3, row=count).value = x['pensionario']
                        ws.cell(column=4, row=count).value = Decimal(
                            x['percent']).quantize(Decimal('0.01'))
                        ws.cell(column=5, row=count).value = (
                            (Decimal(x['remuneracion']) - dom) - sfamily) if x['setfamily'] else (
                                Decimal(x['remuneracion']) - dom)
                        ws.cell(column=6, row=count).value = dom
                        ws.cell(column=7, row=count).value = sfamily
                        ws.cell(column=8, row=count).value = x['remuneracion']
                        ## to here performed calc hours
                        calchour = Decimal((
                            x['remuneracion'] / float(request.GET['payment'])) / pxhour).quantize(
                                Decimal('0.0001'))
                        ws.cell(column=9, row=count).value = x['twork']
                        ws.cell(column=10, row=count).value = x['textra']
                        tworked = Decimal(calchour * x['twork']).quantize(Decimal('0.01'))
                        ws.cell(column=11, row=count).value = tworked + dom
                        safp = Decimal(
                            float(tworked) * (float(x['percent']) / 100)).quantize(Decimal('0.01'))
                        ws.cell(column=12, row=count).value = safp
                        ws.cell(column=13, row=count).value = x['discount']
                        ws.cell(column=14, row=count).value = (x['discount'] + safp)
                        tot = ((tworked + dom) - (safp))
                        ws.cell(column=15, row=count).value = tot
                        total += tot
                        count += 1
                    ws.cell(column=15, row=count).value = total
                if request.GET['tfile'] == 'days':
                    for x in week:
                        dom = 0
                        sfamily = 0
                        dom = Decimal(x['remuneracion'] / float(
                            request.GET['payment'])).quantize(Decimal('0.01'))
                        if x['setfamily'] == True:
                            sfamily = Decimal((float(reminv)/30) * float(request.GET['payment'])).quantize(Decimal('0.01'))
                        ws.cell(column=1, row=count).value = (count - 4)
                        ws.cell(column=2, row=count).value = x['name']
                        ws.cell(column=3, row=count).value = x['pensionario']
                        ws.cell(column=4, row=count).value = Decimal(
                            x['percent']).quantize(Decimal('0.01'))
                        ws.cell(column=5, row=count).value = (
                            (Decimal(x['remuneracion']) - dom) - sfamily) if x['setfamily'] else (
                                Decimal(x['remuneracion']) - dom)
                        ws.cell(column=6, row=count).value = dom
                        ws.cell(column=7, row=count).value = sfamily
                        ws.cell(column=8, row=count).value = x['remuneracion']
                        ## to here performed calc hours
                        calchour = Decimal((
                            x['remuneracion'] / float(request.GET['payment'])) / pxhour).quantize(
                                Decimal('0.0001'))
                        calcday = Decimal((
                            x['remuneracion'] / float(request.GET['payment']))).quantize(
                                Decimal('0.0001'))
                        ws.cell(column=9, row=count).value = x['count']
                        ws.cell(column=10, row=count).value = x['textra']
                        tworked = Decimal((x['count'] * calcday)).quantize(Decimal('0.01'))
                        ws.cell(column=11, row=count).value = tworked + dom
                        safp = Decimal(
                            float(tworked) * (float(x['percent']) / 100)).quantize(Decimal('0.01'))
                        ws.cell(column=12, row=count).value = safp
                        ws.cell(column=13, row=count).value = x['discount']
                        ws.cell(column=14, row=count).value = (x['discount'] + safp)
                        tot = ((tworked + dom) - (safp))
                        ws.cell(column=15, row=count).value = tot
                        total += tot
                        count += 1
                    ws.cell(column=15, row=count).value = total
                style_range(ws, ('A3:O%d' % count), {
                    'border': Border(top=Side(border_style='thin', color=colors.BLACK),
                                    left=Side(border_style='thin', color=colors.BLACK),
                                    bottom=Side(border_style='thin', color=colors.BLACK),
                                    right=Side(border_style='thin', color=colors.BLACK), )},)
                style_range(ws, 'A3:O4', {'bg': PatternFill(
                    start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')})
                style_range(ws, ('O5:O%d' % count), {'bg': PatternFill(
                    start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')})
                ws['A3'].alignment = Alignment(horizontal='center')
                ws.merge_cells('A3:O3')
                ws.merge_cells('A%(n)d:N%(n)d' % {'n':count})
                ws['A%d'%count].value = 'TOTAL'
                ws['A%d'%count].alignment = Alignment(horizontal='right')
                ws.column_dimensions['A'].width = 4
                ws.column_dimensions['B'].width = 35
                ws.column_dimensions['C'].width = 12
                ws.column_dimensions['D'].width = 7
                style_range(ws, ('D5:O%d' % count), {'fnumber': '#,##0.00'})
                # now include employee with payment extra
                count += 3
                second = count
                ws.cell(column=1, row=count).value = 'EXTRAS'
                ws.merge_cells('A%(n)d:G%(n)d'%{'n': count})
                count += 1
                ws.cell(column=1, row=count).value = '#'
                ws.cell(column=2, row=count).value = 'APELLIDOS Y NOMBRES'
                ws.cell(column=3, row=count).value = 'HORAS EXTRAS'
                ws.cell(column=4, row=count).value = 'IMPORTE'
                ws.cell(column=5, row=count).value = 'VIATICOS'
                ws.cell(column=6, row=count).value = 'DSCT'
                ws.cell(column=7, row=count).value = 'TOTAL'
                count += 1
                nb = 1
                totalextra = 0
                tviatical = 0
                thours = 0
                tdiscount = 0
                style_range(ws, ('A%d:G%d' % (count - 2, count - 1)), {
                    'bg': PatternFill(
                        fill_type='solid', start_color='FFFFFF00', end_color='FFFFFF00')})
                for x in week:
                    ws.cell(column=1, row=count).value = (nb)
                    ws.cell(column=2, row=count).value = x['name']
                    ws.cell(column=3, row=count).value = x['textra']
                    calchour = Decimal((
                        x['remuneracion'] / float(request.GET['payment'])) / pxhour).quantize(
                            Decimal('0.0001'))
                    tsc = Decimal(calchour * x['textra']).quantize(Decimal('0.01'))
                    ws.cell(column=4, row=count).value = tsc
                    ws.cell(column=5, row=count).value = x['viatical']
                    ws.cell(column=6, row=count).value = x['discount']
                    ws.cell(column=7, row=count).value = Decimal(
                        (tsc + x['viatical']) - x['discount']).quantize(Decimal('0.01'))
                    totalextra += ws.cell(column=7, row=count).value
                    tviatical += x['viatical']
                    tdiscount += x['discount']
                    thours += x['textra']
                    nb += 1
                    count += 1
                ws.cell(column=3, row=count).value = thours
                ws.cell(column=5, row=count).value = tviatical
                ws.cell(column=6, row=count).value = tdiscount
                ws.cell(column=7, row=count).value = totalextra
                style_range(ws, ('A%d:G%d' % (second, count)), {
                    'border': Border(
                        top=Side(border_style='thin', color=colors.BLACK),
                        left=Side(border_style='thin', color=colors.BLACK),
                        bottom=Side(border_style='thin', color=colors.BLACK),
                        right=Side(border_style='thin', color=colors.BLACK), )})
                wb.save(response)
                return response
            else:
                return render(request, 'rrhh/exportar.html', kwargs)
        except TemplateDoesNotExist as ext:
            raise Http404(ext)

# class delete assistance
class DeleteAssistance(JSONResponseMixin, TemplateView):

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        try:
            kwargs['types'] = TipoEmpleado.objects.filter(flag=True)
            return render(request, 'rrhh/delassistance.html', kwargs)
        except TemplateDoesNotExist as ext:
            raise Http404(ext)

    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        if  request.is_ajax():
            try:
                if 'deleteAssistance' in request.POST:
                    sass = Assistance.objects.filter(
                        assistance__range=(request.POST['start'], request.POST['end']),
                        employee__tipoemple_id=request.POST['type'])
                    for x in sass:
                        x.delete()
                    bass = BalanceAssistance.objects.filter(
                        assistance__range=(request.POST['start'], request.POST['end']),
                        employee__tipoemple_id=request.POST['type'])
                    for x in bass:
                        x.delete()
                    kwargs['status'] = True
            except ObjectDoesNotExist as oex:
                kwargs['status'] = False
                kwargs['raise'] = str(oex)
        return self.render_to_json_response(kwargs)

# #class load data test
class LoadRe(TemplateView):

    def get(self, request, *args, **kwargs):
        import  random
        # import date
        dni = ['10278772','09434564','43504371','47403858','70015965','09279015','19242694','71744452','43572763','44684519','45576052','09164147','43179067','40472915','17589085','41452883','23392439','47720073','10659544','42492201','40894066','43402155','45505146','42661096','00000000','46377134','77498386','44488336','44263241','46296730','42577451','43402156','46279300','47431405','44463959','42333571','72604244','70921414','77070893','70492850','46824555','43414194','42825437','41822143','44322016','40080344','05406738','47542298','43378121','70266206','43316178','44969405','42310463','43984365','40990712','40143637','43598455','40043764','45768794','47999575','71251518','43869036','46879331','48022378','73585896','25798318','44363536']
        count = 0
        if 'pensionary' in request.GET:
            afp = ['RP001', 'RP002', 'RP003', 'RP004', 'RP005']
            name = {'RP001':'HABITAT', 'RP002':'ONP', 'RP003':'PROFUTURO', 'RP004':'PRIMA', 'RP005':'INTEGRA'}
            percent = {'RP001':12.58, 'RP002':13.00, 'RP003': 12.53, 'RP004': 12.58, 'RP005': 12.53}
            count = 0
            for x in dni:
                try:
                    fc = datetime.datetime.today()
                    ds = detPensEmple()
                    ds.empdni_id = x
                    ds.regimenpens_id = afp[random.randrange(5)]
                    ds.fechinicio = fc
                    ds.fechfin = fc.replace(fc.year + 1)
                    ds.cuspp = '%s' % '{:0>20d}'.format(random.randint(10000, 100000))
                    ds.sctr = random.randrange(2)
                    ds.save()
                    count += 1
                except ObjectDoesNotExist as oex:
                    print oex
            kwargs['affected'] = 'LOAD REGIMENS'
        if 'accounts' in request.GET:
            account = None
            count = 0
            for x in dni:
                try:
                    CuentaEmple.objects.get(empdni_id=x)
                except ObjectDoesNotExist as oex:
                    account = '%i-%s-%i' % (random.randint(100, 999), '{:0>8d}'.format(random.randint(100, 9999)), random.randint(100, 999))
                    amount = random.randint(150, 500)
                    ce = CuentaEmple()
                    ce.empdni_id = x
                    ce.cuenta = account
                    ce.tipodepago_id = 'PA002'
                    ce.estado = 'ACTIVO'
                    ce.remuneracion = amount
                    ce.tipocontrato_id = 'CT001'
                    ce.cts = amount
                    ce.gratificacion = amount * 2
                    ce.costxhora = (amount / 7) / 8
                    ce.flag = True
                    ce.save()
                    count+=1
            kwargs['affected'] = 'LOAD ACCOUNTS'
        kwargs['transact'] = count
        return render(request, 'rrhh/test.html', kwargs)
